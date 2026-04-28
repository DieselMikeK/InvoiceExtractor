import os
import tempfile
import unittest

from openpyxl import load_workbook

from invoice_parser import parse_invoice
from spreadsheet_writer import read_spreadsheet_rows, write_invoice_to_spreadsheet


class SpreadsheetWriterTests(unittest.TestCase):
    def test_bill_no_hyperlink_prefers_source_url(self):
        invoice_data = {
            'invoice_number': '743636',
            'vendor': 'S&B Filters',
            'vendor_address': '15461 Slover Avenue, Fontana CA 92337',
            'terms': 'Net 30',
            'date': '4/28/2026',
            'due_date': '5/28/2026',
            'po_number': '0064810',
            'customer': 'Bill Seeberger',
            'total': '386.66',
            'shipping_cost': '19.50',
            'source_path': 'Invoices/invoices_4-28/SB_Order_743636.email.json',
            'source_url': 'https://shopify.com/69841617189/account/orders/token?locale=en-US',
            'line_items': [
                {
                    'item_number': '83-2004',
                    'description': 'Hot Side Intercooler Pipe',
                    'quantity': '1',
                    'unit_price': '166.83',
                    'amount': '166.83',
                }
            ],
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'source_url_test.xlsx')
            write_invoice_to_spreadsheet(output_path, invoice_data)
            wb = load_workbook(output_path)
            link = wb.active.cell(row=2, column=1).hyperlink

        self.assertIsNotNone(link)
        self.assertEqual(link.target, invoice_data['source_url'])

    def test_dpp_discount_exports_blank_sku_only(self):
        invoice_data = {
            'invoice_number': 'INV-1',
            'vendor': 'Test Vendor',
            'vendor_address': '123 Test St',
            'terms': 'Net 30',
            'date': '4/7/2026',
            'due_date': '5/7/2026',
            'po_number': '12345',
            'customer': 'Test Customer',
            'total': '90.00',
            'shipping_cost': '',
            'line_items': [
                {
                    'item_number': 'DPP DISCOUNT',
                    'description': 'Promo Discount',
                    'quantity': '1',
                    'unit_price': '-10.00',
                    'amount': '-10.00',
                    'is_discount': True,
                }
            ],
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'discount_test.xlsx')
            write_invoice_to_spreadsheet(output_path, invoice_data)
            rows = read_spreadsheet_rows(output_path)

        discount_row = next(
            row for row in rows
            if str(row.get('product_service', '')).strip() == 'DPP Discount'
        )
        self.assertEqual(str(discount_row.get('sku', '')).strip(), '')
        self.assertEqual(str(discount_row.get('rate', '')).strip(), '-10.00')
        self.assertEqual(str(discount_row.get('product_service', '')).strip(), 'DPP Discount')

    def test_suspensionmaxx_discount_exports_like_redhead(self):
        pdf_path = os.path.join(
            os.path.dirname(__file__),
            'training',
            'SM',
            'Invoice_260360_DS_from_SuspensionMaxx_Inc.pdf',
        )
        invoice_data = parse_invoice(pdf_path, lambda *args, **kwargs: None)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'suspensionmaxx_discount.xlsx')
            write_invoice_to_spreadsheet(output_path, invoice_data)
            rows = read_spreadsheet_rows(output_path)

        discount_row = next(
            row for row in rows
            if str(row.get('rate', '')).strip() == '-2.37'
        )
        self.assertEqual(str(discount_row.get('type', '')).strip(), 'Item Details')
        self.assertEqual(str(discount_row.get('category', '')).strip(), 'Purchases')
        self.assertEqual(
            str(discount_row.get('product_service', '')).strip(),
            'Inventory Item (Sellable Item)',
        )
        self.assertEqual(str(discount_row.get('sku', '')).strip(), '')

    def test_isspro_discount_exports_like_redhead(self):
        pdf_path = os.path.join(
            os.path.dirname(__file__),
            'training',
            'ISS',
            'Invoice591260-367573-1.PDF',
        )
        invoice_data = parse_invoice(pdf_path, lambda *args, **kwargs: None)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'isspro_discount.xlsx')
            write_invoice_to_spreadsheet(output_path, invoice_data)
            rows = read_spreadsheet_rows(output_path)

        discount_row = next(
            row for row in rows
            if str(row.get('rate', '')).strip() == '-12.51'
        )
        self.assertEqual(str(discount_row.get('type', '')).strip(), 'Item Details')
        self.assertEqual(str(discount_row.get('category', '')).strip(), 'Purchases')
        self.assertEqual(
            str(discount_row.get('product_service', '')).strip(),
            'Inventory Item (Sellable Item)',
        )
        self.assertEqual(str(discount_row.get('sku', '')).strip(), '')


if __name__ == '__main__':
    unittest.main()
