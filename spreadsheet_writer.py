"""Spreadsheet writer: writes parsed invoice data to QuickBooks-compatible CSV/Excel format."""
import csv
import os
import re
from copy import copy
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

try:
    from core_detection import is_core_candidate
except ImportError:
    from app.core_detection import is_core_candidate

try:
    from invoice_parser import get_vendor_default_terms
except ImportError:
    from app.invoice_parser import get_vendor_default_terms

# Alternating row background color
ALT_ROW_FILL = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
SB_DELIVERY_FEE_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
PPE_STOCK_ORDER_FILL = PatternFill(start_color="D8B4FE", end_color="D8B4FE", fill_type="solid")
SKUNEXUS_FAILED_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
MARGIN_LOW_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
SHOPIFY_CORE_MISSING_FILL = PatternFill(start_color="FFFF6666", end_color="FFFF6666", fill_type="solid")
SHOPIFY_CORE_MISMATCH_FILL = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
SB_DELIVERY_FEE_FONT_COLOR = "FFFF0000"
NOT_INVOICE_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
NOT_INVOICE_FILL_DARK = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")


# QuickBooks Bill Import column definitions (matching Taylor's format)
# Columns marked with * are required
COLUMNS = [
    ('bill_no', 'Bill No.'),           # * Invoice number (e.g., I457156)
    ('vendor', '*Vendor'),              # * Required - vendor name (e.g., S&B)
    ('mailing_address', 'Mailing Address'),
    ('terms', 'Terms'),                 # e.g., Net 30
    ('bill_date', '*Bill Date'),        # * Required - invoice date
    ('due_date', 'Due Date'),
    ('location', 'Location'),
    ('memo', 'Memo'),                   # PO number goes here
    ('type', 'Type'),                   # Always "Category Details" for QuickBooks import
    ('category', 'Category/Account'),   # "Purchases" or "Freight and shipping costs"
    ('product_service', 'Product/Service'),
    ('sku', 'SKU'),
    ('qty', 'Qty'),
    ('rate', 'Rate'),
    ('description', 'Description'),     # Product description
    ('amount', 'Amount'),
    ('billable', 'Billable'),
    ('customer_project', 'Customer/Project'),
    ('tax_rate', 'Tax Rate'),
    ('class_field', 'Class'),
    ('skunexus_margin', 'SkuNexus Margin'),
    ('skunexus_validation', 'SkuNexus Validation'),  # Yes/No
    ('skunexus_failed_fields', 'SkuNexus Failed Fields'),  # Which fields failed
    ('shopify_core', 'Shopify CORE'),
    ('duplicate_status', 'Duplicate Status'),
    ('duplicate_reference', 'Duplicate Reference'),
]

# Column indices for validation columns (1-indexed for openpyxl)
COLUMN_INDEX = {key: idx + 1 for idx, (key, _) in enumerate(COLUMNS)}
VALIDATION_COL = COLUMN_INDEX['skunexus_validation']
FAILED_FIELDS_COL = COLUMN_INDEX['skunexus_failed_fields']
MARGIN_COL = COLUMN_INDEX['skunexus_margin']
MARGIN_WARNING_THRESHOLD = 0.1999

PURCHASES_CATEGORY = 'Purchases'
FREIGHT_CATEGORY = 'Freight and shipping costs'
TYPE_CATEGORY = 'Category Details'
TYPE_ITEM = 'Item Details'


def _is_csv(filepath):
    return str(filepath).lower().endswith('.csv')


def _format_export_date(value):
    if value is None or value == '':
        return ''
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime.min.time())
    else:
        text = str(value).strip()
        if not text:
            return ''
        parsed = None
        for fmt in (
            '%m/%d/%Y',
            '%m/%d/%y',
            '%m-%d-%Y',
            '%m-%d-%y',
            '%Y-%m-%d',
            '%Y/%m/%d',
        ):
            try:
                parsed = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if parsed is None:
            return text
    return f"{parsed.month}/{parsed.day}/{parsed.year}"


def _normalize_export_terms(value):
    raw_text = str(value or '').strip()
    text = re.sub(r'\s+', ' ', raw_text)
    if not text:
        return ''

    collapsed = re.sub(r'[^a-z0-9]+', '', text.lower())
    if collapsed == 'dueuponreceipt':
        return 'Due Upon Receipt'

    if collapsed == 'creditcardbulkbill':
        return 'Credit Card'

    if collapsed == 'creditcard':
        return 'Credit Card'

    match = re.fullmatch(r'n(?:et)?(\d+)', collapsed)
    if match:
        return f"Net {int(match.group(1))}"

    match = re.fullmatch(r'n(?:et)?(\d+)th(?:prox)?', collapsed)
    if match:
        return f"Net {int(match.group(1))}th"

    if collapsed == 'net10thprox':
        return 'Net 10th'

    if '\n' in raw_text or '\r' in raw_text or '\t' in raw_text:
        return text
    return raw_text


def _header_for_key(key):
    for column_key, header in COLUMNS:
        if column_key == key:
            return header
    return key


def _header_aliases_for_key(key):
    """Legacy header aliases for backwards compatibility with existing files."""
    if key == 'amount':
        return ['Total Amount']
    return []


def _build_header_map(ws):
    """Return mapping of normalized header text -> 1-based column index."""
    header_map = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is None:
            continue
        key = str(value).strip().lower()
        if key and key not in header_map:
            header_map[key] = col
    return header_map


def _preferred_col_for_key(key):
    return COLUMN_INDEX.get(key)


def _resolve_col_by_key(ws, key, create_if_missing=False):
    """Resolve a worksheet column by key/header, optionally creating it."""
    header = _header_for_key(key)
    header_map = _build_header_map(ws)
    header_candidates = [header] + _header_aliases_for_key(key)
    for candidate in header_candidates:
        header_key = str(candidate).strip().lower()
        if header_key in header_map:
            return header_map[header_key]

    if not create_if_missing:
        return _preferred_col_for_key(key)

    preferred = _preferred_col_for_key(key)
    if preferred and not ws.cell(row=1, column=preferred).value:
        col_idx = preferred
    else:
        col_idx = ws.max_column + 1

    ws.cell(row=1, column=col_idx, value=header)
    ws.cell(row=1, column=col_idx).font = ws.cell(row=1, column=col_idx).font.copy(bold=True)
    return col_idx


def _capture_cell(cell):
    return {
        'value': cell.value,
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment),
        'number_format': cell.number_format,
        'protection': copy(cell.protection),
        'hyperlink': copy(cell.hyperlink) if cell.hyperlink else None,
        'comment': copy(cell.comment) if cell.comment else None,
    }


def _apply_cell_snapshot(cell, snapshot):
    cell.value = snapshot['value']
    cell.font = copy(snapshot['font'])
    cell.fill = copy(snapshot['fill'])
    cell.border = copy(snapshot['border'])
    cell.alignment = copy(snapshot['alignment'])
    cell.number_format = snapshot['number_format']
    cell.protection = copy(snapshot['protection'])
    cell.hyperlink = copy(snapshot['hyperlink']) if snapshot['hyperlink'] else None
    cell.comment = copy(snapshot['comment']) if snapshot['comment'] else None


def _normalize_tail_columns(ws):
    """Normalize tail columns to:
    Class | SkuNexus Margin | SkuNexus Validation | SkuNexus Failed Fields | Shopify CORE | Duplicate Status | Duplicate Reference
    """
    target_cols = {
        'skunexus_margin': 21,
        'skunexus_validation': 22,
        'skunexus_failed_fields': 23,
        'shopify_core': 24,
        'duplicate_status': 25,
        'duplicate_reference': 26,
    }

    header_map = _build_header_map(ws)
    source_cols = {}
    already_aligned = True

    for key, target_col in target_cols.items():
        header = _header_for_key(key)
        source_col = header_map.get(str(header).strip().lower())
        source_cols[key] = source_col
        if source_col != target_col:
            already_aligned = False

    if already_aligned:
        return

    # Ensure target columns exist.
    if ws.max_column < 26:
        ws.cell(row=1, column=26, value=ws.cell(row=1, column=26).value)

    for row in range(1, ws.max_row + 1):
        snapshots = {}
        for key, source_col in source_cols.items():
            if source_col:
                snapshots[key] = _capture_cell(ws.cell(row=row, column=source_col))
            else:
                snapshots[key] = None

        for key, target_col in target_cols.items():
            target_cell = ws.cell(row=row, column=target_col)
            snap = snapshots.get(key)
            if snap is not None:
                _apply_cell_snapshot(target_cell, snap)
            else:
                if row == 1:
                    target_cell.value = _header_for_key(key)
                    target_cell.font = target_cell.font.copy(bold=True)
                else:
                    target_cell.value = ''


def _get_csv_writer(filepath):
    file_exists = os.path.exists(filepath) and os.path.getsize(filepath) > 0
    csv_file = open(filepath, 'a', newline='', encoding='utf-8')
    writer = csv.writer(csv_file)
    if not file_exists:
        writer.writerow([header for _, header in COLUMNS])
    return csv_file, writer


def get_or_create_workbook(filepath):
    """Load existing workbook or create a new one with headers."""
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
        _normalize_tail_columns(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Bills"
        # Write header row
        for col_idx, (key, header) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = cell.font.copy(bold=True)
        # Set reasonable column widths
        widths = {
            'A': 12, 'B': 12, 'C': 25, 'D': 10, 'E': 12,
            'F': 12, 'G': 12, 'H': 15, 'I': 14, 'J': 18,
            'K': 18, 'L': 15, 'M': 6, 'N': 10, 'O': 40, 'P': 12,
            'Q': 10, 'R': 18, 'S': 10, 'T': 12,
            'U': 16, 'V': 18, 'W': 40,  # SkuNexus columns
            'X': 14,  # Shopify CORE
            'Y': 20, 'Z': 40,  # Duplicate columns
        }
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

    return wb, ws


def count_existing_invoice_groups(ws):
    """Count existing invoice groups for alternating row color.

    We count invoice starts, not unique bill numbers. This keeps alternating
    row colors stable even when two invoices share the same Bill No.
    """
    memo_col = _resolve_col_by_key(ws, 'memo', create_if_missing=False)
    mailing_col = _resolve_col_by_key(ws, 'mailing_address', create_if_missing=False)
    terms_col = _resolve_col_by_key(ws, 'terms', create_if_missing=False)
    customer_col = _resolve_col_by_key(ws, 'customer_project', create_if_missing=False)

    count = 0
    saw_any_row = False
    for row in range(2, ws.max_row + 1):  # Skip header row
        bill_cell = ws.cell(row=row, column=1)
        bill_no = str(bill_cell.value or '').strip()
        memo = str(ws.cell(row=row, column=memo_col).value or '').strip() if memo_col else ''
        mailing = str(ws.cell(row=row, column=mailing_col).value or '').strip() if mailing_col else ''
        terms = str(ws.cell(row=row, column=terms_col).value or '').strip() if terms_col else ''
        customer = str(ws.cell(row=row, column=customer_col).value or '').strip() if customer_col else ''
        has_link = bool(getattr(bill_cell, 'hyperlink', None))
        is_invoice_start = has_link or bool(memo) or bool(mailing) or bool(terms) or bool(customer)

        if is_invoice_start:
            count += 1
            saw_any_row = True
        elif not saw_any_row and bill_no:
            # Legacy/fallback rows that may not include the newer markers.
            count += 1
            saw_any_row = True

    return count


def write_invoice_rows(filepath, invoice_data, status_callback=None):
    """Write invoice data as multiple rows (one per line item + shipping).

    Format matches QuickBooks Bill Import:
    - First row: full invoice info with first line item
    - Additional rows: line items with repeated invoice identity fields
    - Shipping row: invoice identity fields + shipping details

    Args:
        filepath: Path to the .xlsx file
        invoice_data: Dict with parsed invoice fields including 'line_items' list
        status_callback: Optional function(msg, tag) for status updates

    Returns:
        int: Number of rows written
    """
    cb = status_callback or (lambda msg, tag=None: None)

    is_csv = _is_csv(filepath)
    wb = None
    ws = None
    csv_file = None
    csv_writer = None
    should_color = False

    if is_csv:
        csv_file, csv_writer = _get_csv_writer(filepath)
    else:
        wb, ws = get_or_create_workbook(filepath)
        # Count existing invoice groups to determine if this invoice should be colored
        # Even-indexed invoices (0, 2, 4...) get no color, odd-indexed (1, 3, 5...) get gray
        invoice_index = count_existing_invoice_groups(ws)
        should_color = (invoice_index % 2 == 1)

    bill_no = invoice_data.get('invoice_number', '')
    vendor = invoice_data.get('vendor', '')
    mailing_address = invoice_data.get('vendor_address', '')
    terms = _normalize_export_terms(
        get_vendor_default_terms(invoice_data.get('vendor', '')) or invoice_data.get('terms', '')
    )
    bill_date = _format_export_date(invoice_data.get('date', ''))
    due_date = _format_export_date(invoice_data.get('due_date', ''))
    memo = invoice_data.get('po_number', '')
    customer = invoice_data.get('customer', '')
    total_amount = invoice_data.get('total', '')
    shared_invoice_fields = {
        'bill_no': bill_no,
        'vendor': vendor,
        'bill_date': bill_date,
        'due_date': due_date,
    }

    line_items = invoice_data.get('line_items', [])
    shipping_cost = invoice_data.get('shipping_cost', '')
    is_stock_order = bool(invoice_data.get('stock_order'))
    stock_order_description = str(invoice_data.get('stock_order_description') or 'STOCK ORDER').strip()

    rows_written = 0

    def _write_row(row_data):
        nonlocal rows_written
        row_num = None
        row_fill = row_data.get('_row_fill')
        is_sb_delivery_fee = bool(row_data.get('_sb_delivery_fee'))
        if row_fill is None and should_color:
            row_fill = ALT_ROW_FILL
        if is_csv:
            csv_writer.writerow([row_data.get(key, '') for key, _ in COLUMNS])
        else:
            header_map = _build_header_map(ws)
            row_num = ws.max_row + 1
            for key, header in COLUMNS:
                col_idx = header_map.get(str(header).strip().lower())
                if not col_idx:
                    col_idx = _resolve_col_by_key(ws, key, create_if_missing=True)
                    header_map = _build_header_map(ws)
                value = row_data.get(key, '')
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                # Apply alternating color per invoice (not per row)
                if row_fill:
                    cell.fill = row_fill
                if is_sb_delivery_fee and key == 'description':
                    cell.font = cell.font.copy(color=SB_DELIVERY_FEE_FONT_COLOR)
        rows_written += 1
        return row_num

    if is_stock_order:
        # PPE stock orders intentionally skip detailed line items and totals.
        stock_context_parts = []
        if bill_no:
            stock_context_parts.append(f"Bill No: {bill_no}")
        if memo:
            stock_context_parts.append(f"Memo (PO): {memo}")
        if stock_context_parts:
            stock_order_description = (
                f"{stock_order_description} | {' | '.join(stock_context_parts)}"
            )

        row_data = {
            'bill_no': bill_no,
            'vendor': vendor,
            'mailing_address': mailing_address,
            'terms': terms,
            'bill_date': bill_date,
            'due_date': due_date,
            'location': '',
            'memo': memo,
            'type': TYPE_CATEGORY,
            'category': '',
            'product_service': '',
            'sku': '',
            'qty': '',
            'rate': '',
            'description': stock_order_description,
            'amount': '',
            'billable': '',
            'customer_project': customer,
            'tax_rate': '',
            'class_field': '',
            '_row_fill': PPE_STOCK_ORDER_FILL,
        }
        first_row_num = _write_row(row_data)
        source_path = invoice_data.get('source_path') or ''
        if (not is_csv) and first_row_num and bill_no and source_path:
            try:
                link_cell = ws.cell(row=first_row_num, column=1)
                link_cell.hyperlink = source_path
            except Exception:
                pass

        if is_csv:
            if csv_file:
                csv_file.close()
        else:
            wb.save(filepath)
        cb(f"  Written {rows_written} row(s) to spreadsheet for invoice {bill_no}", "success")
        return rows_written

    # Write first row with full invoice header + first line item (if any)

    first_item = line_items[0] if line_items else {}

    def _normalize_qty_value(value):
        if value is None:
            return ''
        s = str(value).strip()
        if not s:
            return ''
        try:
            num = float(s.replace(',', ''))
        except (ValueError, TypeError):
            return s
        # If it's effectively an integer, drop decimals
        if abs(num - round(num)) < 1e-9:
            return str(int(round(num)))
        return s

    def _is_discount(item):
        item_num = str(item.get('item_number', '')).lower()
        desc = str(item.get('description', '')).lower()
        return bool(item.get('is_discount')) or ('discount' in item_num) or ('discount' in desc)
    def _item_export_override(item, key):
        value = item.get(key)
        if value is None:
            return None
        text = str(value).strip()
        return text or None
    def _is_core(item):
        item_num = str(item.get('item_number', '')).lower()
        desc = str(item.get('description', '')).lower()
        return is_core_candidate('', item_num, desc)
    def _is_ere(item):
        item_num = str(item.get('item_number', '')).lower().strip()
        desc = str(item.get('description', '')).lower()
        return item_num in ('e.r.e.', 'ere') or 'environmental regulation expense' in desc
    def _normalize_shipping_label(text):
        s = str(text or '').lower()
        if 'drop ship' in s or 'dropship' in s:
            return 'Drop Ship'
        if 'freight' in s or 'frieght' in s:
            return 'Freight'
        if 'ship' in s:
            return 'Shipping'
        return 'Shipping'
    def _row_category_for_item(item, is_freight):
        override = _item_export_override(item, 'qb_category_override')
        if override:
            return override
        if not is_freight:
            return PURCHASES_CATEGORY
        shipping_label = _normalize_shipping_label(item.get('description') or item.get('item_number'))
        if shipping_label == 'Drop Ship':
            return PURCHASES_CATEGORY
        return FREIGHT_CATEGORY
    def _row_type_for_item(item, is_freight):
        override = _item_export_override(item, 'qb_type_override')
        if override:
            return override
        if not is_freight:
            return TYPE_ITEM
        return TYPE_CATEGORY
    def _core_description(item):
        code = str(item.get('item_number', '')).strip()
        desc = str(item.get('description', '')).strip()
        if desc:
            return desc
        return code
    def _product_service_for_item(item, is_discount, is_core, is_ere, is_freight):
        override = _item_export_override(item, 'qb_product_service_override')
        if override:
            return override
        if is_discount:
            return 'DPP Discount'
        if is_freight:
            shipping_label = _normalize_shipping_label(item.get('description') or item.get('item_number'))
            if shipping_label == 'Drop Ship':
                return 'Drop Ship'
            return shipping_label
        return 'Inventory Item (Sellable Item)'
    def _sku_for_item(item):
        override = _item_export_override(item, 'qb_sku_override')
        if override:
            return override
        return str(item.get('item_number', '')).strip()
    def _description_for_item(item, is_discount, is_core, is_freight):
        if is_core:
            return _core_description(item)
        if is_discount:
            return ''
        if is_freight:
            return (str(item.get('description', '')).strip()
                    or str(item.get('item_number', '')).strip())
        return item.get('description', '')

    first_is_freight = bool(first_item.get('is_freight'))
    first_is_discount = _is_discount(first_item)
    first_is_core = _is_core(first_item)
    first_is_ere = _is_ere(first_item)
    first_category = _row_category_for_item(first_item, first_is_freight) if first_item else ''
    first_type = _row_type_for_item(first_item, first_is_freight) if first_item else TYPE_CATEGORY
    if first_item:
        first_product_service = _product_service_for_item(first_item, first_is_discount, first_is_core, first_is_ere, first_is_freight)
        first_sku = _sku_for_item(first_item)
    else:
        first_product_service = ''
        first_sku = ''
    row_data = {
        'bill_no': bill_no,
        'vendor': vendor,
        'mailing_address': mailing_address,
        'terms': terms,
        'bill_date': bill_date,
        'due_date': due_date,
        'location': '',
        'memo': memo,
        'type': first_type,
        'category': first_category if first_item else '',
        'product_service': first_product_service,
        'sku': first_sku,
        'qty': _normalize_qty_value(first_item.get('quantity', '')),
        'rate': first_item.get('unit_price', ''),
        'description': _description_for_item(first_item, first_is_discount, first_is_core, first_is_freight),
        'amount': '',
        'billable': '',
        'customer_project': customer,
        'tax_rate': '',
        'class_field': '',
    }
    if first_item and first_item.get('sb_delivery_fee'):
        row_data['_row_fill'] = SB_DELIVERY_FEE_FILL
        row_data['_sb_delivery_fee'] = True

    first_row_num = _write_row(row_data)
    source_path = invoice_data.get('source_path') or ''
    if (not is_csv) and first_row_num and bill_no and source_path:
        try:
            link_cell = ws.cell(row=first_row_num, column=1)
            link_cell.hyperlink = source_path
        except Exception:
            pass

    # Write additional line items (rows 2+)
    for item in line_items[1:]:
        is_freight = bool(item.get('is_freight'))
        is_discount = _is_discount(item)
        is_core = _is_core(item)
        is_ere = _is_ere(item)
        category = _row_category_for_item(item, is_freight)
        row_data = {
            'bill_no': shared_invoice_fields['bill_no'],
            'vendor': shared_invoice_fields['vendor'],
            'mailing_address': '',
            'terms': '',
            'bill_date': shared_invoice_fields['bill_date'],
            'due_date': shared_invoice_fields['due_date'],
            'location': '',
            'memo': '',
            'type': _row_type_for_item(item, is_freight),
            'category': category,
            'product_service': _product_service_for_item(item, is_discount, is_core, is_ere, is_freight),
            'sku': _sku_for_item(item),
            'qty': _normalize_qty_value(item.get('quantity', '')),
            'rate': item.get('unit_price', ''),
            'description': _description_for_item(item, is_discount, is_core, is_freight),
            'amount': '',
            'billable': '',
            'customer_project': '',
            'tax_rate': '',
            'class_field': '',
        }
        if item.get('sb_delivery_fee'):
            row_data['_row_fill'] = SB_DELIVERY_FEE_FILL
            row_data['_sb_delivery_fee'] = True

        _write_row(row_data)

    # Always write shipping row (even if $0), unless freight items already exist
    has_freight_item = any(item.get('is_freight') for item in line_items)

    try:
        shipping_val = float(str(shipping_cost).replace(',', '').replace('$', ''))
    except (ValueError, TypeError):
        shipping_val = 0

    # Format shipping amount - show 0 if no shipping cost
    shipping_rate = shipping_cost if shipping_val > 0 else '0'
    shipping_desc = invoice_data.get('shipping_description', 'Shipping')
    shipping_label = _normalize_shipping_label(shipping_desc)
    shipping_qty = ''
    if shipping_label == 'Drop Ship':
        shipping_qty = _normalize_qty_value(invoice_data.get('shipping_quantity', ''))
    shipping_category = PURCHASES_CATEGORY if shipping_label == 'Drop Ship' else FREIGHT_CATEGORY
    shipping_type = TYPE_CATEGORY
    shipping_product_service = (
        'Drop Ship' if shipping_label == 'Drop Ship' else shipping_label
    )

    if (not has_freight_item) and (shipping_rate or shipping_desc):
        row_data = {
            'bill_no': shared_invoice_fields['bill_no'],
            'vendor': shared_invoice_fields['vendor'],
            'mailing_address': '',
            'terms': '',
            'bill_date': shared_invoice_fields['bill_date'],
            'due_date': shared_invoice_fields['due_date'],
            'location': '',
            'memo': '',
            'type': shipping_type,
            'category': shipping_category,
            'product_service': shipping_product_service,
            'sku': '',
            'qty': shipping_qty,
            'rate': shipping_rate,
            'description': shipping_desc,
            'amount': '',
            'billable': '',
            'customer_project': '',
            'tax_rate': '',
            'class_field': '',
        }

        _write_row(row_data)

    # Add final total amount row (summary line)
    if total_amount:
        row_data = {
            'bill_no': shared_invoice_fields['bill_no'],
            'vendor': shared_invoice_fields['vendor'],
            'mailing_address': '',
            'terms': '',
            'bill_date': shared_invoice_fields['bill_date'],
            'due_date': shared_invoice_fields['due_date'],
            'location': '',
            'memo': '',
            'type': TYPE_CATEGORY,
            'category': '',
            'product_service': 'Total Amount',
            'sku': '',
            'qty': '',
            'rate': '',
            'description': '',
            'amount': total_amount,
            'billable': '',
            'customer_project': '',
            'tax_rate': '',
            'class_field': '',
        }

        _write_row(row_data)

    if is_csv:
        if csv_file:
            csv_file.close()
    else:
        wb.save(filepath)
    cb(f"  Written {rows_written} row(s) to spreadsheet for invoice {bill_no}", "success")

    return rows_written


# Keep old function for backwards compatibility but redirect to new one
def write_invoice_to_spreadsheet(filepath, invoice_data, status_callback=None):
    """Append invoice data to spreadsheet (wrapper for write_invoice_rows)."""
    return write_invoice_rows(filepath, invoice_data, status_callback)


def write_not_invoice_row(filepath, source_path, status_callback=None):
    """Write a single red 'Not an Invoice' row for a file that failed the invoice check."""
    is_csv = _is_csv(filepath)
    if is_csv:
        return

    wb, ws = get_or_create_workbook(filepath)
    row_num = ws.max_row + 1
    header_map = _build_header_map(ws)

    # Alternate shade if the previous row was also a "Not an Invoice" row
    row_fill = NOT_INVOICE_FILL
    if row_num > 2:
        bill_col = header_map.get('bill no.') or COLUMN_INDEX.get('bill_no', 1)
        prev_cell = ws.cell(row=row_num - 1, column=bill_col)
        if str(prev_cell.value or '').strip() == 'Not an Invoice':
            prev_rgb = str(prev_cell.fill.start_color.rgb if prev_cell.fill else '').upper()
            # If previous row was light red, use dark red; otherwise use light red
            row_fill = NOT_INVOICE_FILL_DARK if prev_rgb.endswith('FFCCCC') else NOT_INVOICE_FILL

    for key, header in COLUMNS:
        col_idx = header_map.get(str(header).strip().lower())
        if not col_idx:
            col_idx = _resolve_col_by_key(ws, key, create_if_missing=True)
            header_map = _build_header_map(ws)
        value = 'Not an Invoice' if key == 'bill_no' else ''
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.fill = row_fill

    # Hyperlink "Not an Invoice" in Bill No. column to the source file
    bill_col = header_map.get('bill no.')
    if not bill_col:
        bill_col = COLUMN_INDEX.get('bill_no', 1)
    if source_path:
        try:
            ws.cell(row=row_num, column=bill_col).hyperlink = source_path
        except Exception:
            pass

    wb.save(filepath)


def read_spreadsheet_rows(filepath):
    """Read all data rows from the spreadsheet.

    Returns:
        list: List of dicts, one per row, with column keys from COLUMNS
    """
    if not os.path.exists(filepath):
        return []

    if _is_csv(filepath):
        rows = []
        with open(filepath, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for idx, row in enumerate(reader, start=2):  # Header is row 1
                row_data = {'_row_num': idx}
                for key, header in COLUMNS:
                    row_data[key] = row.get(header, '') or ''
                rows.append(row_data)
        return rows

    wb = load_workbook(filepath)
    ws = wb.active
    header_map = _build_header_map(ws)

    rows = []
    for row_num in range(2, ws.max_row + 1):  # Skip header
        row_data = {'_row_num': row_num}
        for key, header in COLUMNS:
            col_idx = header_map.get(str(header).strip().lower())
            if not col_idx:
                col_idx = _preferred_col_for_key(key)
            row_data[key] = ws.cell(row=row_num, column=col_idx).value or ''
        rows.append(row_data)

    return rows


def _ensure_validation_headers(ws):
    """Ensure validation/margin headers exist and return their column indices."""
    margin_col = _resolve_col_by_key(ws, 'skunexus_margin', create_if_missing=True)
    validation_col = _resolve_col_by_key(ws, 'skunexus_validation', create_if_missing=True)
    failed_col = _resolve_col_by_key(ws, 'skunexus_failed_fields', create_if_missing=True)
    shopify_core_col = _resolve_col_by_key(ws, 'shopify_core', create_if_missing=True)
    return margin_col, validation_col, failed_col, shopify_core_col


def _set_cell_horizontal_alignment(cell, horizontal):
    align = copy(cell.alignment)
    align.horizontal = horizontal
    cell.alignment = align


def _normalize_validation_alignment(ws, validation_col):
    for row_num in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=validation_col)
        value = cell.value
        if value is None or str(value).strip() == '':
            continue
        _set_cell_horizontal_alignment(cell, 'center')


def _normalize_margin_alignment(ws, margin_col):
    for row_num in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=margin_col)
        value = cell.value
        if value is None or str(value).strip() == '':
            continue
        _set_cell_horizontal_alignment(cell, 'center')


def _normalize_shopify_core_alignment(ws, shopify_core_col):
    for row_num in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=shopify_core_col)
        value = cell.value
        if value is None or str(value).strip() == '':
            continue
        _set_cell_horizontal_alignment(cell, 'center')


def _apply_validation_to_ws(ws, row_num, validation_col, failed_col, is_valid, failed_fields):
    """Apply validation values to a single row in an open worksheet."""
    validation_cell = ws.cell(row=row_num, column=validation_col)
    failed_cell = ws.cell(row=row_num, column=failed_col)
    if is_valid is None:
        validation_cell.value = ''
        failed_cell.value = ''
    else:
        validation_cell.value = 'Yes' if is_valid else 'No'
        failed_cell.value = ', '.join(failed_fields) if failed_fields else ''
    _set_cell_horizontal_alignment(validation_cell, 'center')

    # Apply alternating color to validation cells (match existing row color)
    first_cell = ws.cell(row=row_num, column=1)
    if first_cell.fill and first_cell.fill.patternType == 'solid':
        row_fill = PatternFill(
            start_color=first_cell.fill.start_color.rgb,
            end_color=first_cell.fill.end_color.rgb,
            fill_type='solid'
        )
        validation_cell.fill = row_fill
        failed_cell.fill = row_fill

    if is_valid is False:
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=col_idx).fill = SKUNEXUS_FAILED_FILL


def _apply_margin_to_ws(ws, row_num, margin_col, margin_value):
    """Apply margin value and warning fill to one worksheet row."""
    margin_cell = ws.cell(row=row_num, column=margin_col)
    _set_cell_horizontal_alignment(margin_cell, 'center')

    # Default fill follows row pattern.
    first_cell = ws.cell(row=row_num, column=1)
    if first_cell.fill and first_cell.fill.patternType == 'solid':
        row_fill = PatternFill(
            start_color=first_cell.fill.start_color.rgb,
            end_color=first_cell.fill.end_color.rgb,
            fill_type='solid'
        )
        margin_cell.fill = row_fill

    if margin_value is None or margin_value == '':
        margin_cell.value = ''
        return

    try:
        margin_num = float(margin_value)
    except (ValueError, TypeError):
        margin_cell.value = str(margin_value)
        return

    margin_cell.value = round(margin_num, 4)
    if margin_num < MARGIN_WARNING_THRESHOLD:
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=col_idx).fill = MARGIN_LOW_FILL


def _parse_shopify_core_update(core_update):
    if isinstance(core_update, dict):
        value = core_update.get('value', '')
        status = str(core_update.get('status', '')).strip().lower()
        return value, status
    if isinstance(core_update, (tuple, list)):
        if len(core_update) >= 2:
            return core_update[0], str(core_update[1] or '').strip().lower()
        if len(core_update) == 1:
            return core_update[0], ''
    return core_update, ''


def _apply_shopify_core_to_ws(ws, row_num, shopify_core_col, core_update):
    core_value, status = _parse_shopify_core_update(core_update)
    core_cell = ws.cell(row=row_num, column=shopify_core_col)
    _set_cell_horizontal_alignment(core_cell, 'center')

    # Default fill follows row pattern.
    first_cell = ws.cell(row=row_num, column=1)
    if first_cell.fill and first_cell.fill.patternType == 'solid':
        row_fill = PatternFill(
            start_color=first_cell.fill.start_color.rgb,
            end_color=first_cell.fill.end_color.rgb,
            fill_type='solid'
        )
        core_cell.fill = row_fill

    if core_value is None or str(core_value).strip() == '':
        core_cell.value = ''
    else:
        core_cell.value = str(core_value)

    # Row-level warning fills for Shopify CORE checks.
    fill = None
    if status == 'missing':
        fill = SHOPIFY_CORE_MISSING_FILL
    elif status == 'mismatch':
        fill = SHOPIFY_CORE_MISMATCH_FILL
    if fill:
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=col_idx).fill = fill


def _write_validation_results_csv(filepath, updates, margin_updates=None, shopify_core_updates=None):
    if not updates or not os.path.exists(filepath):
        # Allow margin-only updates.
        if (not margin_updates and not shopify_core_updates) or not os.path.exists(filepath):
            return

    with open(filepath, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        rows = list(reader)

    if not headers:
        headers = [header for _, header in COLUMNS]

    for header in ('SkuNexus Validation', 'SkuNexus Failed Fields', 'SkuNexus Margin', 'Shopify CORE'):
        if header not in headers:
            headers.append(header)

    for row_num, (is_valid, failed_fields) in (updates or {}).items():
        idx = row_num - 2  # Convert 1-indexed row number (header is row 1)
        if idx < 0 or idx >= len(rows):
            continue
        if is_valid is None:
            validation_value = ''
            failed_value = ''
        else:
            validation_value = 'Yes' if is_valid else 'No'
            failed_value = ', '.join(failed_fields) if failed_fields else ''
        rows[idx]['SkuNexus Validation'] = validation_value
        rows[idx]['SkuNexus Failed Fields'] = failed_value

    for row_num, margin_value in (margin_updates or {}).items():
        idx = row_num - 2
        if idx < 0 or idx >= len(rows):
            continue
        if margin_value is None or margin_value == '':
            rows[idx]['SkuNexus Margin'] = ''
            continue
        try:
            rows[idx]['SkuNexus Margin'] = f"{float(margin_value):.4f}"
        except (ValueError, TypeError):
            rows[idx]['SkuNexus Margin'] = str(margin_value)

    for row_num, core_update in (shopify_core_updates or {}).items():
        idx = row_num - 2
        if idx < 0 or idx >= len(rows):
            continue
        core_value, _ = _parse_shopify_core_update(core_update)
        if core_value is None:
            rows[idx]['Shopify CORE'] = ''
        else:
            rows[idx]['Shopify CORE'] = str(core_value)

    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({h: row.get(h, '') for h in headers})


def _write_validation_results_xlsx(filepath, updates, margin_updates=None, shopify_core_updates=None):
    if not updates and not margin_updates and not shopify_core_updates:
        return
    wb = load_workbook(filepath)
    ws = wb.active
    _normalize_tail_columns(ws)
    margin_col, validation_col, failed_col, shopify_core_col = _ensure_validation_headers(ws)
    for row_num, margin_value in (margin_updates or {}).items():
        _apply_margin_to_ws(ws, row_num, margin_col, margin_value)
    for row_num, (is_valid, failed_fields) in (updates or {}).items():
        _apply_validation_to_ws(ws, row_num, validation_col, failed_col, is_valid, failed_fields)
    for row_num, core_update in (shopify_core_updates or {}).items():
        _apply_shopify_core_to_ws(ws, row_num, shopify_core_col, core_update)
    _normalize_validation_alignment(ws, validation_col)
    _normalize_margin_alignment(ws, margin_col)
    _normalize_shopify_core_alignment(ws, shopify_core_col)
    try:
        wb.save(filepath)
    except PermissionError as e:
        raise PermissionError(
            f"{filepath} is locked (likely open in Excel). "
            "Close the file and run validation again."
        ) from e


def write_validation_results(filepath, updates, margin_updates=None, shopify_core_updates=None):
    """Write multiple validation results in one pass.

    Args:
        filepath: Path to the spreadsheet
        updates: dict {row_num: (is_valid, failed_fields)}
        margin_updates: optional dict {row_num: margin_float}
        shopify_core_updates: optional dict {row_num: (value, status)}
    """
    if _is_csv(filepath):
        _write_validation_results_csv(filepath, updates, margin_updates, shopify_core_updates)
    else:
        _write_validation_results_xlsx(filepath, updates, margin_updates, shopify_core_updates)


def write_validation_result(filepath, row_num, is_valid, failed_fields, margin_value=None):
    """Write validation result to a specific row.

    Args:
        filepath: Path to the spreadsheet
        row_num: The row number to update (1-indexed, header is row 1)
        is_valid: Boolean - True for Yes, False for No
        failed_fields: List of field names that failed validation
        margin_value: Optional SkuNexus margin value for this row
    """
    margin_updates = None
    if margin_value is not None:
        margin_updates = {row_num: margin_value}
    write_validation_results(filepath, {row_num: (is_valid, failed_fields)}, margin_updates)


def get_unique_po_numbers(filepath):
    """Get unique PO numbers from the spreadsheet.

    Returns:
        dict: {po_number: [list of row_nums with that PO]}
    """
    rows = read_spreadsheet_rows(filepath)
    po_rows = {}

    for row in rows:
        memo = row.get('memo', '')
        if memo:
            # Normalize PO number
            po_num = str(memo).strip()
            if po_num not in po_rows:
                po_rows[po_num] = []
            po_rows[po_num].append(row)

    return po_rows
