import os
import tempfile
import unittest
from datetime import datetime, timezone

from openpyxl import load_workbook

from invoice_extractor_gui import (
    _build_date_range_time_filter,
    _build_date_range_time_query,
    InvoiceExtractorGUI,
    _build_today_time_query,
    _cell_fill_rgb,
    _format_time_value,
    _get_status_messages,
    _is_diamond_eye_zero_shipping_batch_row,
    _load_sender_sidecar,
    _merge_sender_metadata_entries,
    _parse_time_input,
    _lookup_sender_metadata_entry,
    _should_preserve_duplicate_row_fill,
    _save_sender_sidecar,
)
from spreadsheet_writer import write_invoice_to_spreadsheet


class DiamondEyeBatchExportFilterTests(unittest.TestCase):
    def test_drops_zero_shipping_row_for_diamond_eye(self):
        row = {
            'vendor': 'Diamond Eye Manufacturing - $3.00 DS Fee',
            'category': 'Freight and shipping costs',
            'product_service': 'Shipping',
            'sku': '',
            'rate': '0',
        }

        self.assertTrue(_is_diamond_eye_zero_shipping_batch_row(row))

    def test_keeps_positive_shipping_row_for_diamond_eye(self):
        row = {
            'vendor': 'Diamond Eye Manufacturing',
            'category': 'Freight and shipping costs',
            'product_service': 'Shipping',
            'sku': '',
            'rate': '18.75',
        }

        self.assertFalse(_is_diamond_eye_zero_shipping_batch_row(row))

    def test_keeps_zero_shipping_row_for_other_vendors(self):
        row = {
            'vendor': 'ATS Diesel',
            'category': 'Freight and shipping costs',
            'product_service': 'Shipping',
            'sku': '',
            'rate': '0',
        }

        self.assertFalse(_is_diamond_eye_zero_shipping_batch_row(row))

class SenderMetadataLookupTests(unittest.TestCase):
    def test_falls_back_to_filename_when_source_path_misses(self):
        entries = {
            'oldroot/Invoices/Invoice_123.pdf': {
                'source_file': 'oldroot/Invoices/Invoice_123.pdf',
                'filename': 'Invoice_123.pdf',
                'sender_email': 'invoicing@kcturbos.com',
            }
        }

        entry = _lookup_sender_metadata_entry(
            entries,
            'newroot/Invoices/Invoice_123.pdf',
            'Invoice_123.pdf',
        )

        self.assertEqual(entry.get('sender_email'), 'invoicing@kcturbos.com')

    def test_exact_path_entry_is_enriched_by_richer_filename_match(self):
        entries = {
            'Invoices/invoices_4-16/Invoice_123.pdf': {
                'source_file': 'Invoices/invoices_4-16/Invoice_123.pdf',
                'filename': 'Invoice_123.pdf',
                'sender_email': 'noreply@suspension.randysww.com',
                'subject': 'Invoice attached',
                'message_text': '',
            },
            'Invoices/invoices_4-16_2/Invoice_123.pdf': {
                'source_file': 'Invoices/invoices_4-16_2/Invoice_123.pdf',
                'filename': 'Invoice_123.pdf',
                'sender_email': 'noreply@suspension.randysww.com',
                'subject': 'Invoice attached',
                'message_text': 'Sincerely,\nCarli Suspension',
            },
        }

        entry = _lookup_sender_metadata_entry(
            entries,
            'Invoices/invoices_4-16/Invoice_123.pdf',
            'Invoice_123.pdf',
        )

        self.assertEqual(entry.get('source_file'), 'Invoices/invoices_4-16/Invoice_123.pdf')
        self.assertEqual(entry.get('message_text'), 'Sincerely,\nCarli Suspension')

    def test_sender_sidecar_round_trip(self):
        entry = {
            'source_file': 'Invoices/Invoice_123.pdf',
            'filename': 'Invoice_123.pdf',
            'sender_email': 'invoicing@kcturbos.com',
            'sender_header': 'KC Turbos <invoicing@kcturbos.com>',
            'subject': 'Invoice attached',
            'message_text': 'Forwarded message body',
            'message_id': 'abc123',
            'downloaded_at': '2026-04-07 14:00:00',
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            pdf_path = os.path.join(tmpdir, 'Invoice_123.pdf')
            with open(pdf_path, 'wb') as f:
                f.write(b'%PDF-1.4')
            _save_sender_sidecar(pdf_path, entry)
            loaded = _load_sender_sidecar(pdf_path)

        self.assertEqual(loaded.get('sender_email'), 'invoicing@kcturbos.com')
        self.assertEqual(loaded.get('message_id'), 'abc123')
        self.assertEqual(loaded.get('message_text'), 'Forwarded message body')

    def test_merge_sender_metadata_entries_fills_stale_sidecar_fields(self):
        sidecar_entry = {
            'source_file': 'Invoices/invoices_4-16/Invoice_123.pdf',
            'filename': 'Invoice_123.pdf',
            'sender_email': 'noreply@suspension.randysww.com',
            'sender_header': '<noreply@suspension.randysww.com>',
            'subject': 'Invoice attached',
            'message_text': '',
            'message_id': 'old-message',
            'downloaded_at': '2026-04-16 11:52:35',
        }
        cached_entry = {
            'source_file': 'Invoices/invoices_4-16_2/Invoice_123.pdf',
            'filename': 'Invoice_123.pdf',
            'sender_email': 'noreply@suspension.randysww.com',
            'sender_header': '<noreply@suspension.randysww.com>',
            'subject': 'Invoice attached',
            'message_text': 'Sincerely,\nCarli Suspension',
            'message_id': 'new-message',
            'downloaded_at': '2026-04-16 14:38:36',
        }

        merged = _merge_sender_metadata_entries(sidecar_entry, cached_entry)

        self.assertEqual(merged.get('source_file'), 'Invoices/invoices_4-16/Invoice_123.pdf')
        self.assertEqual(merged.get('message_text'), 'Sincerely,\nCarli Suspension')
        self.assertEqual(merged.get('message_id'), 'old-message')


class GmailTodayTimeQueryTests(unittest.TestCase):
    def test_parse_time_input_accepts_24_hour_and_ampm(self):
        self.assertEqual(_parse_time_input('14:35').strftime('%H:%M'), '14:35')
        self.assertEqual(_parse_time_input('2:35 PM').strftime('%H:%M'), '14:35')

    def test_format_time_value_uses_compact_ampm_display(self):
        self.assertEqual(_format_time_value(_parse_time_input('2:35 PM')), '2:35 PM')

    def test_build_today_time_query_before_uses_start_of_day_to_boundary(self):
        reference = datetime(2026, 4, 7, 15, 0, tzinfo=timezone.utc)

        query = _build_today_time_query('10:30', 'Before', reference)

        expected_start = int(datetime(2026, 4, 7, 0, 0, tzinfo=timezone.utc).timestamp())
        expected_boundary = int(datetime(2026, 4, 7, 10, 30, tzinfo=timezone.utc).timestamp())
        self.assertEqual(query, f"after:{expected_start} before:{expected_boundary}")

    def test_build_today_time_query_after_uses_boundary_to_end_of_day(self):
        reference = datetime(2026, 4, 7, 15, 0, tzinfo=timezone.utc)

        query = _build_today_time_query('2:30 PM', 'After', reference)

        expected_boundary = int(datetime(2026, 4, 7, 14, 30, tzinfo=timezone.utc).timestamp())
        expected_end = int(datetime(2026, 4, 8, 0, 0, tzinfo=timezone.utc).timestamp())
        self.assertEqual(query, f"after:{expected_boundary} before:{expected_end}")

    def test_build_today_time_query_rejects_invalid_time(self):
        reference = datetime(2026, 4, 7, 15, 0, tzinfo=timezone.utc)

        with self.assertRaises(ValueError):
            _build_today_time_query('bad-time', 'Before', reference)


class GmailDateRangeTimeQueryTests(unittest.TestCase):
    def test_build_date_range_time_query_uses_explicit_bounds(self):
        reference = datetime(2026, 4, 7, 15, 0, tzinfo=timezone.utc)

        query = _build_date_range_time_query(
            datetime(2026, 4, 7, tzinfo=timezone.utc).date(),
            datetime(2026, 4, 8, tzinfo=timezone.utc).date(),
            from_time_value='9:15 AM',
            to_time_value='2:45 PM',
            reference_dt=reference,
        )

        expected_start = int(datetime(2026, 4, 7, 9, 15, tzinfo=timezone.utc).timestamp())
        expected_end = int(datetime(2026, 4, 8, 14, 46, tzinfo=timezone.utc).timestamp())
        self.assertEqual(query, f"after:{expected_start} before:{expected_end}")

    def test_build_date_range_time_filter_defaults_to_whole_day_when_times_blank(self):
        reference = datetime(2026, 4, 7, 15, 0, tzinfo=timezone.utc)

        time_filter = _build_date_range_time_filter(
            datetime(2026, 4, 7, tzinfo=timezone.utc).date(),
            datetime(2026, 4, 7, tzinfo=timezone.utc).date(),
            reference_dt=reference,
        )

        expected_start = int(datetime(2026, 4, 7, 0, 0, tzinfo=timezone.utc).timestamp())
        expected_end = int(datetime(2026, 4, 8, 0, 0, tzinfo=timezone.utc).timestamp())
        self.assertEqual(time_filter, {'start_ts': expected_start, 'end_ts': expected_end})

    def test_build_date_range_time_filter_rejects_reversed_datetime_bounds(self):
        reference = datetime(2026, 4, 7, 15, 0, tzinfo=timezone.utc)

        with self.assertRaises(ValueError):
            _build_date_range_time_filter(
                datetime(2026, 4, 8, tzinfo=timezone.utc).date(),
                datetime(2026, 4, 8, tzinfo=timezone.utc).date(),
                from_time_value='3:00 PM',
                to_time_value='2:00 PM',
                reference_dt=reference,
            )

    def test_build_date_range_time_filter_rejects_time_without_matching_date(self):
        reference = datetime(2026, 4, 7, 15, 0, tzinfo=timezone.utc)

        with self.assertRaises(ValueError):
            _build_date_range_time_filter(
                None,
                datetime(2026, 4, 8, tzinfo=timezone.utc).date(),
                from_time_value='3:00 PM',
                reference_dt=reference,
            )


class StatusMessageTests(unittest.TestCase):
    def test_does_not_include_extra_test_status_line(self):
        missing_required_dir = os.path.join(
            os.getcwd(),
            '__missing_required_dir_for_status_test__',
        )
        messages = _get_status_messages(missing_required_dir, ocr_available=True)

        self.assertEqual(messages[0]['text'], 'Connected')
        self.assertEqual(messages[0]['foreground'], 'green')
        self.assertNotIn('test', [message['text'] for message in messages])


class DuplicateHighlightTests(unittest.TestCase):
    def test_preserves_stock_order_fill_when_duplicate_marked(self):
        invoice_data = {
            'invoice_number': 'STOCK-1',
            'vendor': 'Turn 14 Distribution',
            'vendor_address': '100 Tournament Dr. Horsham, PA 19044',
            'terms': 'Credit Card',
            'date': '4/10/2026',
            'due_date': '',
            'po_number': '0058197',
            'customer': 'Diesel Power Products',
            'total': '',
            'shipping_cost': '',
            'stock_order': True,
            'stock_order_description': 'STOCK ORDER',
            'source_path': r'Invoices\\Turn14_Invoice_16039445.pdf',
            'line_items': [],
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'duplicates.xlsx')
            write_invoice_to_spreadsheet(output_path, invoice_data)

            gui = InvoiceExtractorGUI.__new__(InvoiceExtractorGUI)
            gui._apply_duplicate_flags(
                output_path,
                history_by_po={
                    '0058197': [
                        {
                            'po_number': '0058197',
                            'source_file': 'Invoices/Turn14_SalesOrder_16271604.pdf',
                            'downloaded_at': '2026-04-08 10:00:00',
                        }
                    ]
                },
                history_by_bill={},
            )

            ws = load_workbook(output_path).active
            first_cell = ws.cell(row=2, column=1)
            self.assertTrue(_should_preserve_duplicate_row_fill(first_cell))
            self.assertTrue(_cell_fill_rgb(first_cell).endswith('D8B4FE'))
            self.assertEqual(
                ws.cell(row=2, column=25).value,
                'Duplicate PO (history)',
            )


if __name__ == '__main__':
    unittest.main()
