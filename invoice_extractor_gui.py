#!/usr/bin/env python3
"""
Invoice Extractor - GUI Application
Connects to Gmail, downloads invoice attachments, parses them,
and exports extracted data to an Excel spreadsheet.
"""

import os
import sys
import json
import csv
import re
import threading
import tkinter as tk
from tkinter import messagebox, ttk
import tkinter.font as tkfont
import time
import ctypes
import shutil
import subprocess
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False
try:
    from tkcalendar import Calendar
    TKCALENDAR_AVAILABLE = True
except Exception:
    TKCALENDAR_AVAILABLE = False

from gmail_client import (
    GmailClient,
    DriveHistoryClient,
    PROCESSED_LABEL_NAME,
    WrongAuthorizedAccountError
)
try:
    from core_detection import is_core_candidate
except ImportError:
    from app.core_detection import is_core_candidate
from invoice_parser import parse_invoice, OCR_AVAILABLE
from spreadsheet_writer import (
    COLUMNS, write_invoice_to_spreadsheet, write_not_invoice_row,
    read_spreadsheet_rows,
    write_validation_result, write_validation_results, get_unique_po_numbers
)
from skunexus_client import SkuNexusClient, validate_po_row
from shopify_client import ShopifyClient
from update_utils import (
    MAIN_EXECUTABLE_NAME,
    fetch_release_manifest,
    load_app_version,
    parse_version_tuple,
    stage_release_manifest,
    stage_updater_executable,
)

# Batch export settings (easy to change)
# QuickBooks import appears to cap CSV files at ~100 total lines.
# Keep each batch at 100 lines max including the header row.
BATCH_TOTAL_LINE_LIMIT = 100
BATCH_ROW_LIMIT = BATCH_TOTAL_LINE_LIMIT - 1  # Max data rows per CSV batch
BATCH_FOLDER_PREFIX = "Batch_"
BATCHES_ROOT_NAME = "Batches"
AUTHORIZED_GMAIL_ACCOUNT = "dppautoap@gmail.com"
SHOPIFY_CORE_RATE_TOLERANCE = 0.01
SENDER_METADATA_FIELDNAMES = [
    'source_file',
    'filename',
    'sender_email',
    'sender_header',
    'subject',
    'message_id',
    'downloaded_at',
]

def _normalize_vendor_key(name):
    if not name:
        return ''
    s = name.lower().strip()
    s = s.replace('&', 'and')
    s = ''.join(ch for ch in s if ch.isalnum())
    return s


def _split_vendor_aliases(value):
    if not value:
        return []
    parts = re.split(r'[|;]', str(value))
    return [p.strip() for p in parts if p.strip()]


def load_vendor_aliases(preferred_dir):
    """Load vendor alias map from vendors.csv, preferring external files over the bundled copy."""
    candidates = [
        os.path.join(preferred_dir, 'vendors.csv'),
        os.path.join(os.path.dirname(preferred_dir), 'vendors.csv'),
        get_resource_path('vendors.csv'),
    ]
    path = next((candidate for candidate in candidates if os.path.exists(candidate)), '')
    if not path:
        return {}
    try:
        with open(path, newline='', encoding='utf-8') as f:
            rows = list(csv.reader(f))
    except Exception:
        return {}
    if not rows:
        return {}

    header = [str(c).strip().lower() for c in rows[0]]
    has_header = any(
        h in ('vendor', 'invoice_vendor', 'aliases', 'alias', 'additional_names')
        for h in header
    )
    if not has_header:
        return {}

    def col(row, *names):
        for name in names:
            if name in header:
                idx = header.index(name)
                if idx < len(row):
                    return str(row[idx]).strip()
        return ''

    aliases = {}
    for row in rows[1:]:
        if not row:
            continue
        vendor = col(row, 'vendor')
        if not vendor:
            continue
        alias_value = col(row, 'aliases', 'alias', 'additional_names', 'invoice_vendor')
        if not alias_value and 'skunexus_vendor' in header:
            alias_value = col(row, 'skunexus_vendor')
        alias_list = _split_vendor_aliases(alias_value)
        if not alias_list:
            continue
        key = _normalize_vendor_key(vendor)
        if not key:
            continue
        bucket = aliases.setdefault(key, [])
        for alias in alias_list:
            if alias and alias not in bucket:
                bucket.append(alias)

    return aliases


def _looks_like_sku(value):
    if value is None:
        return False
    raw = str(value).strip()
    if not raw:
        return False
    lower = raw.lower().strip()
    # Explicit non-SKU labels / summary rows
    if 'core' in lower:
        return False
    normalized = re.sub(r'[^a-z0-9]+', '', lower)
    if not normalized:
        return False
    if normalized in {
        'core', 'ere', 'dppdiscount', 'discount',
        'dropship', 'shipping', 'freight',
        'totalamount', 'total', 'subtotal',
    }:
        return False
    # Require at least one digit to avoid matching plain words
    if not any(ch.isdigit() for ch in normalized):
        return False
    # Avoid very short tokens
    return len(normalized) >= 3


def _get_row_sku(row):
    sku = str(row.get('sku', '')).strip()
    if sku:
        return sku
    return str(row.get('product_service', '')).strip()


def _to_float_value(value):
    if value is None:
        return None
    try:
        text = str(value).strip().replace(',', '').replace('$', '')
        if text == '':
            return None
        return float(text)
    except (TypeError, ValueError):
        return None


def _extract_related_order_numbers(sn_data):
    numbers = []
    seen = set()
    if not isinstance(sn_data, dict):
        return numbers

    related_rows = sn_data.get('allRelatedOrders') or []
    if not related_rows:
        one = sn_data.get('relatedOrder') or {}
        if one:
            related_rows = [one]

    for row in related_rows:
        label = str((row or {}).get('label', '')).strip()
        if not label:
            continue
        digits = ''.join(ch for ch in label if ch.isdigit())
        normalized = (digits.lstrip('0') or '0') if digits else label
        if normalized in seen:
            continue
        seen.add(normalized)
        numbers.append(normalized)

    return numbers


def _is_core_row(row):
    row_type = str(row.get('type', '')).strip().lower()
    if row_type != 'item details':
        return False
    category = str(row.get('category', '')).strip().lower()
    if category != 'purchases':
        return False

    product_service = str(row.get('product_service', '')).strip().lower()
    sku = str(row.get('sku', '')).strip().lower()
    description = str(row.get('description', '')).strip().lower()
    return is_core_candidate(product_service, sku, description)


def get_base_dir():
    """Get the base directory - works for both script and PyInstaller exe."""
    if getattr(sys, 'frozen', False):
        # Running as PyInstaller bundle - use exe's directory
        return os.path.dirname(sys.executable)
    else:
        # Running as script
        return os.path.dirname(os.path.abspath(__file__))


def get_resource_path(relative_path):
    """Get path to bundled resource (works for PyInstaller onefile)."""
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)


def _extract_date_tag_from_filename(filename):
    match = re.search(r'(?:invoices_output_|Invoices_Master_)(\d{1,2}-\d{1,2})', filename)
    if match:
        return match.group(1)
    return None


_single_instance_mutex = None


def _bring_existing_window_to_front(window_title):
    if os.name != 'nt':
        return False
    try:
        user32 = ctypes.windll.user32
        hwnd = user32.FindWindowW(None, window_title)
        if hwnd:
            SW_RESTORE = 9
            user32.ShowWindow(hwnd, SW_RESTORE)
            user32.SetForegroundWindow(hwnd)
            return True
    except Exception:
        pass
    return False


def _ensure_single_instance(window_title):
    if os.name != 'nt':
        return True
    try:
        kernel32 = ctypes.windll.kernel32
        mutex = kernel32.CreateMutexW(None, False, "InvoiceExtractor_SingleInstance")
        already_exists = kernel32.GetLastError() == 183  # ERROR_ALREADY_EXISTS
        if already_exists:
            _bring_existing_window_to_front(window_title)
            return False
        global _single_instance_mutex
        _single_instance_mutex = mutex
    except Exception:
        # If mutex fails for some reason, allow app to run.
        return True
    return True


class InvoiceExtractorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Invoice Extractor")
        self._set_window_icon()
        self.root.geometry("750x650")
        self.root.resizable(True, True)

        self.base_dir = get_base_dir()
        if getattr(sys, 'frozen', False):
            app_dir_candidates = (
                os.path.join(self.base_dir, 'App'),
                os.path.join(self.base_dir, 'app'),
            )
            self.app_dir = next(
                (path for path in app_dir_candidates if os.path.isdir(path)),
                app_dir_candidates[0]
            )
        else:
            # Running from source inside App folder
            self.app_dir = self.base_dir
        self.required_dir = os.path.join(self.app_dir, 'required')
        os.makedirs(self.app_dir, exist_ok=True)
        os.makedirs(self.required_dir, exist_ok=True)
        self._sync_runtime_app_files()
        self.invoices_root = os.path.join(self.base_dir, 'Invoices')
        self.batches_root = os.path.join(self.base_dir, BATCHES_ROOT_NAME)
        os.makedirs(self.invoices_root, exist_ok=True)
        os.makedirs(self.batches_root, exist_ok=True)
        self._migrate_required_files()
        self.output_file, self.invoices_dir = self._get_next_run_paths()

        self.is_running = False
        self.header_label = None
        self.header_image = None
        self.header_src_image = None
        self.header_src_width = None
        self.header_base_width = None
        self.header_current_width = None
        self.header_path = None
        self.header_animating = False
        self.header_shrunken = False
        self._header_anim_token = None
        self.header_base_left = 0
        self.last_batches_dir = None
        self._calendar_open = False
        self._calendar_warned = False
        self._calendar_suppress_widget = None
        self._calendar_suppress_until = 0.0
        self.app_version = load_app_version()
        self.available_update = None
        self.update_button = None
        self.update_button_glow = None
        self._update_button_visible = False
        self._update_flash_job = None
        self._update_flash_on = False
        self._update_button_bg = '#157347'
        self._update_button_active_bg = '#198754'
        self._update_button_disabled_bg = '#4f6f5a'
        self._update_button_glow_dim = '#2d7a42'
        self._update_button_glow_bright = '#5dff8a'
        self._update_button_neutral_bg = ttk.Style().lookup('TFrame', 'background') or self.root.cget('bg')
        self.version_var = tk.StringVar(value=f"v{self.app_version}")

        self.build_ui()
        self.root.after(2000, self._check_for_updates_async)

    def _get_update_target_exe_path(self):
        """Return the executable path that should be replaced during updates."""
        if getattr(sys, 'frozen', False):
            return sys.executable

        candidates = [
            os.path.join(self.base_dir, 'dist', MAIN_EXECUTABLE_NAME),
            os.path.join(self.base_dir, MAIN_EXECUTABLE_NAME),
            os.path.join(os.path.dirname(self.base_dir), MAIN_EXECUTABLE_NAME),
        ]
        for candidate in candidates:
            if os.path.exists(candidate):
                return candidate
        return ''

    def _check_for_updates_async(self):
        """Check the remote release manifest without blocking the UI."""
        if not self._get_update_target_exe_path():
            return
        threading.Thread(target=self._check_for_updates, daemon=True).start()

    def _check_for_updates(self):
        """Load update metadata and surface the update button when needed."""
        try:
            manifest = fetch_release_manifest(self.required_dir)
        except Exception as exc:
            # Network failures are expected sometimes; malformed manifests are worth logging.
            if isinstance(exc, (ValueError, json.JSONDecodeError)):
                self.root.after(0, lambda: self.log(f"Update check failed: {exc}", "warning"))
            return

        has_download = bool(str(manifest.get('download_url') or '').strip())
        current_tuple = parse_version_tuple(self.app_version)
        latest_tuple = parse_version_tuple(manifest.get('version'))
        if latest_tuple > current_tuple and has_download:
            self.root.after(0, lambda: self._set_available_update(manifest))
            return

        if latest_tuple > current_tuple and not has_download:
            self.root.after(
                0,
                lambda: self.log(
                    "Update manifest found a newer version but no download URL was provided.",
                    "warning",
                )
            )
        self.root.after(0, lambda: self._set_available_update(None))

    def _set_available_update(self, manifest):
        """Show or hide the update button based on the available release manifest."""
        self.available_update = manifest
        if self.update_button is None or self.update_button_glow is None:
            return

        if manifest:
            self.update_button.configure(text=f"Update to v{manifest['version']}")
            if not self._update_button_visible:
                self.update_button_glow.pack(side=tk.RIGHT, padx=(0, 8))
                self._update_button_visible = True
        elif self._update_button_visible:
            self._stop_update_button_flash()
            self.update_button_glow.pack_forget()
            self._update_button_visible = False

        self._refresh_update_button_state()

    def _stop_update_button_flash(self, glow_color=None):
        """Stop the update pulse and leave the glow in a stable state."""
        if self._update_flash_job is not None:
            self.root.after_cancel(self._update_flash_job)
            self._update_flash_job = None
        self._update_flash_on = False
        if self.update_button_glow is not None:
            self.update_button_glow.configure(bg=glow_color or self._update_button_neutral_bg)

    def _pulse_update_button(self):
        """Pulse the update glow while a clickable update is available."""
        if (
            self.update_button is None
            or self.update_button_glow is None
            or not self.available_update
            or not self._update_button_visible
            or str(self.update_button.cget('state')) != tk.NORMAL
        ):
            self._stop_update_button_flash()
            return

        self._update_flash_on = not self._update_flash_on
        glow_color = (
            self._update_button_glow_bright
            if self._update_flash_on
            else self._update_button_glow_dim
        )
        self.update_button_glow.configure(bg=glow_color)
        self._update_flash_job = self.root.after(550, self._pulse_update_button)

    def _start_update_button_flash(self):
        """Start pulsing the update button if it is not already animating."""
        if self._update_flash_job is not None:
            return
        self._pulse_update_button()

    def _refresh_update_button_state(self):
        """Enable the update button only when an update is available and the app is idle."""
        if self.update_button is None or self.update_button_glow is None:
            return
        if not self.available_update or not self._update_button_visible:
            self._stop_update_button_flash()
            return
        is_enabled = not self.is_running
        state = tk.NORMAL if is_enabled else tk.DISABLED
        self.update_button.configure(state=state)
        if is_enabled:
            self.update_button.configure(
                bg=self._update_button_bg,
                activebackground=self._update_button_active_bg,
                cursor='hand2',
            )
            self._start_update_button_flash()
        else:
            self.update_button.configure(
                bg=self._update_button_disabled_bg,
                activebackground=self._update_button_disabled_bg,
                cursor='arrow',
            )
            self._stop_update_button_flash(glow_color=self._update_button_glow_dim)

    def _on_update_clicked(self):
        """Confirm and launch the external updater helper."""
        if not self.available_update:
            return
        if self.is_running:
            messagebox.showwarning(
                "Update Unavailable",
                "Stop the current job before running the updater.",
            )
            return

        manifest = self.available_update
        lines = [
            f"Install Invoice Extractor v{manifest['version']} now?",
            f"Current version: v{self.app_version}",
        ]
        notes = str(manifest.get('notes') or '').strip()
        if notes:
            lines.extend(['', 'Release notes:', notes])

        if not messagebox.askyesno("Update Available", "\n".join(lines), parent=self.root):
            return

        target_exe = self._get_update_target_exe_path()
        if not target_exe:
            messagebox.showerror(
                "Update Failed",
                "Could not determine which InvoiceExtractor.exe should be updated.",
                parent=self.root,
            )
            return

        try:
            updater_exe = stage_updater_executable(
                manifest.get('version') or self.app_version,
                manifest=manifest,
            )
            manifest_path = stage_release_manifest(manifest, manifest.get('version'))
        except Exception as exc:
            messagebox.showerror(
                "Update Failed",
                f"Could not prepare the updater files.\n\n{exc}",
                parent=self.root,
            )
            return

        primary_download_url = str(manifest.get('download_url') or '').strip()
        args = [
            updater_exe,
            '--current-exe', target_exe,
            '--manifest-file', manifest_path,
            '--target-version', manifest['version'],
            '--source-version', self.app_version,
            '--wait-pid', str(os.getpid()),
        ]
        if primary_download_url:
            args.extend(['--download-url', primary_download_url])
        expected_hash = str(manifest.get('sha256') or '').strip()
        if expected_hash:
            args.extend(['--sha256', expected_hash])

        try:
            self.update_button.configure(state=tk.DISABLED)
            subprocess.Popen(args, cwd=os.path.dirname(target_exe) or None)
        except Exception as exc:
            self._refresh_update_button_state()
            messagebox.showerror(
                "Update Failed",
                f"Could not launch the updater helper.\n\n{exc}",
                parent=self.root,
            )
            return

        self.root.destroy()

    def _get_next_run_paths(self):
        """Pick the next available output file and invoices folder (same suffix)."""
        date_tag = f"{datetime.now().month}-{datetime.now().day}"
        output_base = f"Invoices_Master_{date_tag}"
        folder_base = f"invoices_{date_tag}"
        ext = '.xlsx'

        output_path = os.path.join(self.base_dir, f"{output_base}{ext}")
        folder_path = os.path.join(self.invoices_root, folder_base)
        if (not os.path.exists(output_path)) and (not os.path.exists(folder_path)):
            return output_path, folder_path

        for i in range(2, 10000):
            output_candidate = os.path.join(self.base_dir, f"{output_base}_{i}{ext}")
            folder_candidate = os.path.join(self.invoices_root, f"{folder_base}_{i}")
            if (not os.path.exists(output_candidate)) and (not os.path.exists(folder_candidate)):
                return output_candidate, folder_candidate

        # Fallback: use timestamp if we somehow hit a huge count
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join(self.base_dir, f"{output_base}_{stamp}{ext}")
        folder_path = os.path.join(self.invoices_root, f"{folder_base}_{stamp}")
        return output_path, folder_path

    def _get_output_files_for_validation(self):
        """Return all output XLSX files in base_dir (sorted oldest to newest)."""
        files = []
        for name in os.listdir(self.base_dir):
            lower = name.lower()
            if not lower.endswith('.xlsx'):
                continue
            if not (lower.startswith('invoices_output') or lower.startswith('invoices_master_')):
                continue
            files.append(os.path.join(self.base_dir, name))
        files.sort(key=lambda p: os.path.getmtime(p))
        return files

    def _find_master_spreadsheets(self):
        """Return all master invoice XLSX files in base_dir (newest first)."""
        files = []
        for name in os.listdir(self.base_dir):
            lower = name.lower()
            if not lower.endswith('.xlsx'):
                continue
            if not (lower.startswith('invoices_output') or lower.startswith('invoices_master_')):
                continue
            if name.startswith('~$'):
                continue
            files.append(os.path.join(self.base_dir, name))
        files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return files

    def _select_master_for_batching(self):
        """Pick the best master spreadsheet for batching (prefer today's date tag)."""
        masters = self._find_master_spreadsheets()
        if not masters:
            return None
        date_tag = f"{datetime.now().month}-{datetime.now().day}"
        today_matches = [
            p for p in masters
            if f"Invoices_Master_{date_tag}" in os.path.basename(p)
        ]
        if today_matches:
            return today_matches[0]
        return masters[0]

    def _get_next_batches_dir(self, date_tag):
        base_name = f"{BATCH_FOLDER_PREFIX}{date_tag}"
        candidate = os.path.join(self.batches_root, base_name)
        if not os.path.exists(candidate):
            return candidate
        for i in range(2, 10000):
            candidate = os.path.join(self.batches_root, f"{base_name}_{i}")
            if not os.path.exists(candidate):
                return candidate
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return os.path.join(self.batches_root, f"{base_name}_{stamp}")

    def _migrate_required_files(self):
        files = [
            'client_secret.json',
            'skunexus_config.json',
            'shopify_config.json',
            'shopify_token.json',
            'invoice_history.csv',
        ]
        # Do not auto-migrate token.pickle. If a user deletes it, they expect
        # the next run to force a fresh OAuth login.
        for name in files:
            dest = os.path.join(self.required_dir, name)
            if os.path.exists(dest):
                continue
            legacy_app_dir = os.path.join(self.base_dir, 'app')
            for src_dir in (self.app_dir, legacy_app_dir, self.base_dir):
                src = os.path.join(src_dir, name)
                if os.path.exists(src):
                    try:
                        os.replace(src, dest)
                    except Exception:
                        try:
                            shutil.copy2(src, dest)
                        except Exception:
                            pass
                    break

    def _sync_runtime_app_files(self):
        """Refresh curated runtime files that should track the shipped application version."""
        for name in ('vendors.csv',):
            source = get_resource_path(name)
            if not os.path.exists(source):
                continue
            destination = os.path.join(self.app_dir, name)
            if os.path.abspath(source) == os.path.abspath(destination):
                continue
            try:
                shutil.copy2(source, destination)
            except Exception:
                pass

    def _history_log_path(self):
        return os.path.join(self.required_dir, 'invoice_history.csv')

    def _sender_metadata_path(self):
        return os.path.join(self.required_dir, 'invoice_sender_metadata.csv')

    def _load_sender_metadata(self):
        """Load locally cached sender metadata keyed by source_file."""
        path = self._sender_metadata_path()
        if not os.path.exists(path):
            return {}
        entries = {}
        try:
            with open(path, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    cleaned = {
                        key: str((row or {}).get(key, '') or '').strip()
                        for key in SENDER_METADATA_FIELDNAMES
                    }
                    source_file = cleaned.get('source_file', '')
                    if source_file:
                        entries[source_file] = cleaned
        except Exception as e:
            self.log(f"Warning: could not read sender metadata ({e})", "warning")
        return entries

    def _save_sender_metadata(self, entries):
        """Persist sender metadata locally for future re-parses."""
        path = self._sender_metadata_path()
        try:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=SENDER_METADATA_FIELDNAMES)
                writer.writeheader()
                for entry in entries:
                    writer.writerow({
                        key: str((entry or {}).get(key, '') or '').strip()
                        for key in SENDER_METADATA_FIELDNAMES
                    })
        except Exception as e:
            self.log(f"Warning: could not save sender metadata ({e})", "warning")

    def _load_invoice_history(self, drive_client=None):
        """Load invoice history from Drive (preferred) or local fallback."""
        if drive_client is not None:
            rows = drive_client.download_rows()
            if rows:
                self._save_local_history(rows)
                return rows
        # Local fallback
        path = self._history_log_path()
        if not os.path.exists(path):
            return []
        rows = []
        try:
            with open(path, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    rows.append({k: (v or '').strip() for k, v in row.items()})
        except Exception as e:
            self.log(f"Warning: could not read invoice history ({e})", "warning")
        return rows

    def _save_local_history(self, rows):
        """Write rows to local invoice_history.csv as a cache."""
        from gmail_client import HISTORY_FIELDNAMES
        path = self._history_log_path()
        try:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=HISTORY_FIELDNAMES)
                writer.writeheader()
                for row in rows:
                    writer.writerow({k: row.get(k, '') for k in HISTORY_FIELDNAMES})
        except Exception as e:
            self.log(f"Warning: could not cache invoice history locally ({e})", "warning")

    def _append_invoice_history(self, entries, drive_client=None):
        """Append new entries to Drive history (preferred) and local fallback."""
        if not entries:
            return
        from gmail_client import HISTORY_FIELDNAMES
        if drive_client is not None:
            # Download fresh, merge, upload
            existing = drive_client.download_rows()
            existing.extend(entries)
            drive_client.upload_rows(existing)
            self._save_local_history(existing)
            return
        # Local fallback
        path = self._history_log_path()
        file_exists = os.path.exists(path)
        try:
            with open(path, 'a', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=HISTORY_FIELDNAMES)
                if not file_exists:
                    writer.writeheader()
                for entry in entries:
                    writer.writerow({k: entry.get(k, '') for k in HISTORY_FIELDNAMES})
        except Exception as e:
            self.log(f"Warning: could not update invoice history ({e})", "warning")

    def _apply_duplicate_flags(self, filepath, history_by_po, history_by_bill=None):
        """Mark duplicates in output spreadsheet and return duplicate summary."""
        if not os.path.exists(filepath):
            return {'duplicate_invoices': 0, 'duplicate_rows': 0}
        if str(filepath).lower().endswith('.csv'):
            return {'duplicate_invoices': 0, 'duplicate_rows': 0}

        history_by_po = history_by_po or {}
        history_by_bill = history_by_bill or {}

        wb = load_workbook(filepath)
        ws = wb.active
        if ws.max_row < 2:
            return {'duplicate_invoices': 0, 'duplicate_rows': 0}

        header_map = {}
        for col in range(1, ws.max_column + 1):
            header_val = ws.cell(row=1, column=col).value
            if header_val is None:
                continue
            header_key = str(header_val).strip().lower()
            if header_key and header_key not in header_map:
                header_map[header_key] = col

        col_keys = [key for key, _ in COLUMNS]

        def _col_for(header_text, key_name):
            header_key = str(header_text).strip().lower()
            if header_key in header_map:
                return header_map[header_key]
            try:
                return col_keys.index(key_name) + 1
            except ValueError:
                return None

        dup_status_col = _col_for('Duplicate Status', 'duplicate_status')
        dup_ref_col = _col_for('Duplicate Reference', 'duplicate_reference')
        memo_col = _col_for('Memo', 'memo')
        mailing_col = _col_for('Mailing Address', 'mailing_address')
        terms_col = _col_for('Terms', 'terms')
        customer_col = _col_for('Customer/Project', 'customer_project')

        if not all([dup_status_col, dup_ref_col, memo_col, mailing_col, terms_col, customer_col]):
            return {'duplicate_invoices': 0, 'duplicate_rows': 0}

        # Build explicit invoice entries so duplicate Bill No. invoices remain separate.
        entries = []
        row_to_entry = {}
        current_entry_idx = None
        prev_bill = ''

        for row_num in range(2, ws.max_row + 1):
            has_any_value = False
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row_num, column=col).value
                if val is not None and str(val).strip():
                    has_any_value = True
                    break
            if not has_any_value:
                continue

            bill_cell = ws.cell(row=row_num, column=1)
            bill_no = str(bill_cell.value or '').strip()
            memo = str(ws.cell(row=row_num, column=memo_col).value or '').strip()
            mailing = str(ws.cell(row=row_num, column=mailing_col).value or '').strip()
            terms = str(ws.cell(row=row_num, column=terms_col).value or '').strip()
            customer = str(ws.cell(row=row_num, column=customer_col).value or '').strip()
            has_link = bool(getattr(bill_cell, 'hyperlink', None))

            # First row of an invoice always has at least one header-like marker.
            start_marker = has_link or bool(memo) or bool(mailing) or bool(terms) or bool(customer)
            bill_changed = bool(bill_no and prev_bill and bill_no != prev_bill)
            starts_new = (current_entry_idx is None) or start_marker or bill_changed

            if starts_new:
                source_ref = ''
                if has_link and getattr(bill_cell, 'hyperlink', None):
                    try:
                        source_ref = str(bill_cell.hyperlink.target or '').strip()
                    except Exception:
                        source_ref = ''
                entries.append({
                    'index': len(entries),
                    'start_row': row_num,
                    'bill_no': bill_no,
                    'po': memo,
                    'rows': [],
                    'source_ref': source_ref,
                })
                current_entry_idx = len(entries) - 1

            entry = entries[current_entry_idx]
            if bill_no and not entry['bill_no']:
                entry['bill_no'] = bill_no
            if memo and not entry['po']:
                entry['po'] = memo
            if has_link and getattr(bill_cell, 'hyperlink', None) and not entry.get('source_ref'):
                try:
                    entry['source_ref'] = str(bill_cell.hyperlink.target or '').strip()
                except Exception:
                    pass
            entry['rows'].append(row_num)
            row_to_entry[row_num] = current_entry_idx

            if bill_no:
                prev_bill = bill_no

        if not entries:
            return {'duplicate_invoices': 0, 'duplicate_rows': 0}

        bill_to_entries = {}
        po_to_entries = {}
        for entry in entries:
            bill_no = entry.get('bill_no', '')
            po = entry.get('po', '')
            if bill_no:
                bill_to_entries.setdefault(bill_no, []).append(entry['index'])
            if po:
                po_to_entries.setdefault(po, []).append(entry['index'])

        status_by_entry = {}
        ref_by_entry = {}

        def _short_source_label(path_value):
            value = str(path_value or '').strip()
            if not value:
                return ''
            value = value.replace('\\', '/')
            base = os.path.basename(value)
            return base or value

        def _history_ref(hist_row):
            src = _short_source_label(hist_row.get('source_file', ''))
            date_val = str(hist_row.get('invoice_date', '')).strip()
            if not date_val:
                downloaded = str(hist_row.get('downloaded_at', '')).strip()
                if downloaded:
                    date_val = downloaded.split(' ')[0]
            if src and date_val:
                return f"{src} ({date_val})"
            if src:
                return src
            if date_val:
                return date_val
            bill = str(hist_row.get('bill_no', '')).strip()
            po_number = str(hist_row.get('po_number', '')).strip()
            if bill and po_number:
                return f"{bill} / {po_number}"
            return bill or po_number or ''

        for entry in entries:
            idx = entry['index']
            bill_no = entry.get('bill_no', '')
            po = entry.get('po', '')
            status_parts = []
            ref_parts = []

            same_bill_entries = bill_to_entries.get(bill_no, []) if bill_no else []
            same_po_entries = po_to_entries.get(po, []) if po else []
            same_invoice_entries = []
            if bill_no and po:
                same_invoice_entries = [
                    other_idx for other_idx in same_bill_entries
                    if other_idx in set(same_po_entries)
                ]

            if len(same_invoice_entries) > 1:
                status_parts.append("Duplicate Invoice (current file)")
                current_refs = []
                seen = set()
                for other_idx in same_invoice_entries:
                    if other_idx == idx:
                        continue
                    other = entries[other_idx]
                    ref = _short_source_label(other.get('source_ref')) or (
                        (other.get('bill_no') or '') + (f" / {other.get('po')}" if other.get('po') else '')
                    ).strip()
                    if ref and ref not in seen:
                        current_refs.append(ref)
                        seen.add(ref)
                if current_refs:
                    ref_parts.append("Current file: " + ", ".join(current_refs[:5]))
            else:
                if len(same_bill_entries) > 1:
                    status_parts.append("Duplicate Bill No. (current file)")
                    current_refs = []
                    seen = set()
                    for other_idx in same_bill_entries:
                        if other_idx == idx:
                            continue
                        other = entries[other_idx]
                        ref = _short_source_label(other.get('source_ref')) or str(other.get('po') or '').strip()
                        if ref and ref not in seen:
                            current_refs.append(ref)
                            seen.add(ref)
                    if current_refs:
                        ref_parts.append("Current Bill No.: " + ", ".join(current_refs[:5]))

                if len(same_po_entries) > 1:
                    status_parts.append("Duplicate PO (current file)")
                    current_refs = []
                    seen = set()
                    for other_idx in same_po_entries:
                        if other_idx == idx:
                            continue
                        other = entries[other_idx]
                        ref = _short_source_label(other.get('source_ref')) or str(other.get('bill_no') or '').strip()
                        if ref and ref not in seen:
                            current_refs.append(ref)
                            seen.add(ref)
                    if current_refs:
                        ref_parts.append("Current PO: " + ", ".join(current_refs[:5]))

            history_po_entries = history_by_po.get(po, []) if po else []
            history_bill_entries = history_by_bill.get(bill_no, []) if bill_no else []
            history_same_invoice = []
            if bill_no and po and history_bill_entries:
                for hist in history_bill_entries:
                    hist_po = str(hist.get('po_number', '')).strip()
                    if hist_po and hist_po == po:
                        history_same_invoice.append(hist)

            if history_same_invoice:
                status_parts.append("Duplicate Invoice (history)")
                hist_refs = []
                seen_refs = set()
                for hist in history_same_invoice:
                    ref = _history_ref(hist)
                    if ref and ref not in seen_refs:
                        hist_refs.append(ref)
                        seen_refs.add(ref)
                if hist_refs:
                    ref_parts.append("History: " + ", ".join(hist_refs[:5]))
            else:
                if history_po_entries:
                    status_parts.append("Duplicate PO (history)")
                    hist_refs = []
                    seen_refs = set()
                    for hist in history_po_entries:
                        ref = _history_ref(hist)
                        if ref and ref not in seen_refs:
                            hist_refs.append(ref)
                            seen_refs.add(ref)
                    if hist_refs:
                        ref_parts.append("History PO: " + ", ".join(hist_refs[:5]))

                if history_bill_entries:
                    status_parts.append("Duplicate Bill No. (history)")
                    hist_refs = []
                    seen_refs = set()
                    for hist in history_bill_entries:
                        ref = _history_ref(hist)
                        if ref and ref not in seen_refs:
                            hist_refs.append(ref)
                            seen_refs.add(ref)
                    if hist_refs:
                        ref_parts.append("History Bill No.: " + ", ".join(hist_refs[:5]))

            if status_parts:
                unique_status = []
                seen_status = set()
                for part in status_parts:
                    if part not in seen_status:
                        unique_status.append(part)
                        seen_status.add(part)
                unique_refs = []
                seen_refs = set()
                for part in ref_parts:
                    if part and part not in seen_refs:
                        unique_refs.append(part)
                        seen_refs.add(part)
                status_by_entry[idx] = "; ".join(unique_status)
                ref_by_entry[idx] = " | ".join(unique_refs)

        dup_entry_indices = set(status_by_entry.keys())
        dup_fill_light = PatternFill(start_color="A8A8A8", end_color="A8A8A8", fill_type='solid')
        dup_fill_dark = PatternFill(start_color="888888", end_color="888888", fill_type='solid')

        for row_num in range(2, ws.max_row + 1):
            entry_idx = row_to_entry.get(row_num)
            if entry_idx is None:
                continue

            entry = entries[entry_idx]
            is_first = entry.get('start_row') == row_num
            status = status_by_entry.get(entry_idx)
            if is_first and status:
                ws.cell(row=row_num, column=dup_status_col, value=status)
                ws.cell(row=row_num, column=dup_ref_col, value=ref_by_entry.get(entry_idx, ''))
            else:
                ws.cell(row=row_num, column=dup_status_col, value='')
                ws.cell(row=row_num, column=dup_ref_col, value='')

            # Match row fill for duplicate columns first.
            first_cell = ws.cell(row=row_num, column=1)
            is_yellow = False
            if first_cell.fill and first_cell.fill.patternType == 'solid':
                color = first_cell.fill.start_color.rgb or first_cell.fill.start_color.index
                if color and str(color).upper().endswith('FFFF00'):
                    is_yellow = True
                row_fill = PatternFill(
                    start_color=first_cell.fill.start_color.rgb,
                    end_color=first_cell.fill.end_color.rgb,
                    fill_type='solid'
                )
                ws.cell(row=row_num, column=dup_status_col).fill = row_fill
                ws.cell(row=row_num, column=dup_ref_col).fill = row_fill

            # Override whole row for duplicate entries with darker alternating gray.
            if entry_idx in dup_entry_indices and not is_yellow:
                dup_fill = dup_fill_dark if (entry_idx % 2 == 0) else dup_fill_light
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_num, column=col).fill = dup_fill

        wb.save(filepath)

        duplicate_rows = sum(len(entries[idx]['rows']) for idx in dup_entry_indices)
        return {
            'duplicate_invoices': len(dup_entry_indices),
            'duplicate_rows': duplicate_rows,
        }

    def _refresh_batch_buttons(self):
        masters = self._find_master_spreadsheets()
        state = tk.NORMAL if masters else tk.DISABLED
        if hasattr(self, 'export_batches_button'):
            self.export_batches_button.config(state=state)
        if hasattr(self, 'validate_button'):
            outputs = self._get_output_files_for_validation()
            validate_state = tk.NORMAL if outputs else tk.DISABLED
            self.validate_button.config(state=validate_state)

    def _set_window_icon(self):
        """Set the window/taskbar icon (Tk default is the leaf)."""
        try:
            icon_path = get_resource_path('logo.ico')
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except Exception:
            # If icon can't be set (e.g., missing file), keep default.
            pass

    def _open_date_picker(self, target_var):
        return self._open_date_picker_at(target_var, None)

    def _open_date_picker_at(self, target_var, anchor_widget, title=None):
        if not TKCALENDAR_AVAILABLE:
            if not getattr(self, '_calendar_warned', False):
                self.log("Calendar picker unavailable (tkcalendar not installed).", "warning")
                self._calendar_warned = True
            return
        if getattr(self, '_calendar_open', False):
            return
        # Suppress immediate reopen after close (regardless of widget)
        if time.time() < getattr(self, '_calendar_suppress_until', 0):
            return

        self._calendar_open = True
        top = tk.Toplevel(self.root)
        top.title(title or "Select Date")
        top.transient(self.root)
        top.grab_set()

        cal = Calendar(top, selectmode='day', date_pattern='yyyy/mm/dd')
        # Preselect if existing
        existing = self._parse_date_input(target_var.get().strip())
        if existing:
            cal.selection_set(existing)
        cal.pack(padx=10, pady=10)

        def _close():
            self._calendar_open = False
            self._calendar_suppress_widget = anchor_widget
            self._calendar_suppress_until = time.time() + 0.5
            top.grab_release()
            top.destroy()

        def _set_date():
            target_var.set(cal.get_date())
            _close()

        btn_frame = ttk.Frame(top)
        btn_frame.pack(pady=(0, 10))
        ttk.Button(btn_frame, text="Set Date", command=_set_date).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_frame, text="Cancel", command=_close).pack(side=tk.LEFT)
        top.protocol("WM_DELETE_WINDOW", _close)

        # Position popup above the input (fallback to below if needed)
        top.update_idletasks()
        if anchor_widget is not None:
            x = anchor_widget.winfo_rootx()
            y = anchor_widget.winfo_rooty() - top.winfo_height() - 8
            if y < 0:
                y = anchor_widget.winfo_rooty() + anchor_widget.winfo_height() + 8
            top.geometry(f"+{x}+{y}")

    def _parse_date_input(self, value):
        if not value:
            return None
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        return None

    def _build_gmail_query(self):
        """Build Gmail search query based on date filter options."""
        if self.today_filter_var.get():
            today = datetime.now().date()
            after = today.strftime("%Y/%m/%d")
            before = (today + timedelta(days=1)).strftime("%Y/%m/%d")
            return f"after:{after} before:{before}", "today"

        if self.date_filter_var.get():
            from_raw = self.date_from_var.get().strip()
            to_raw = self.date_to_var.get().strip()
            if not from_raw and not to_raw:
                self.log("Date filter is enabled but no dates were provided.", "error")
                return None, None
            from_date = self._parse_date_input(from_raw)
            to_date = self._parse_date_input(to_raw)
            if from_raw and not from_date:
                self.log("Invalid From date. Use YYYY/MM/DD.", "error")
                return None, None
            if to_raw and not to_date:
                self.log("Invalid To date. Use YYYY/MM/DD.", "error")
                return None, None
            if from_date and to_date and from_date > to_date:
                self.log("From date must be on or before To date.", "error")
                return None, None

            parts = []
            if from_date:
                parts.append(f"after:{from_date.strftime('%Y/%m/%d')}")
            if to_date:
                inclusive_before = to_date + timedelta(days=1)
                parts.append(f"before:{inclusive_before.strftime('%Y/%m/%d')}")
            return " ".join(parts), "range"

        return f"-label:{PROCESSED_LABEL_NAME}", "label"

    def _update_date_filter_state(self):
        enabled = self.date_filter_var.get()
        state = tk.NORMAL if enabled else tk.DISABLED
        if hasattr(self, 'date_from_entry'):
            self.date_from_entry.config(state=state)
        if hasattr(self, 'date_to_entry'):
            self.date_to_entry.config(state=state)

    def _on_date_filter_toggle(self):
        if self.date_filter_var.get():
            self.today_filter_var.set(False)
        self._update_date_filter_state()

    def _on_today_filter_toggle(self):
        if self.today_filter_var.get():
            self.date_filter_var.set(False)
        self._update_date_filter_state()

    def build_ui(self):
        """Build the main application UI."""
        # Main container with padding
        main_frame = ttk.Frame(self.root, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)

        top_bar = ttk.Frame(main_frame)
        top_bar.pack(fill=tk.X, pady=(0, 2))

        version_label = ttk.Label(
            top_bar,
            textvariable=self.version_var,
            foreground='#888888'
        )
        version_label.pack(side=tk.RIGHT)

        self.update_button_glow = tk.Frame(
            top_bar,
            bg=self._update_button_neutral_bg,
            bd=0,
            highlightthickness=0,
        )

        self.update_button = tk.Button(
            self.update_button_glow,
            text="Update Available",
            command=self._on_update_clicked,
            font=('Segoe UI', 9, 'bold'),
            bg=self._update_button_bg,
            activebackground=self._update_button_active_bg,
            fg='white',
            activeforeground='white',
            disabledforeground='#e2efe6',
            relief=tk.FLAT,
            bd=0,
            padx=14,
            pady=6,
            highlightthickness=0,
            cursor='hand2',
        )
        self.update_button.pack(padx=2, pady=2)

        # Header image (centered)
        self.header_path = get_resource_path('header.png')
        if os.path.exists(self.header_path):
            try:
                self.root.update_idletasks()
                window_width = self.root.winfo_width()
                if window_width <= 1:
                    window_width = 750
                # Base header size (80% width), then reduce by 35%
                target_width = int(window_width * 0.8 * 0.65)
                self.header_image = self._load_header_image(target_width)
                self.header_label = ttk.Label(main_frame, image=self.header_image, anchor='center')
                base_width = self.header_image.width() if self.header_image else None
                self.header_base_left = 0
                self.header_label.pack(fill=tk.X, pady=(0, 5))
                self.header_base_width = base_width
                self.header_current_width = self.header_base_width
            except Exception:
                self.header_image = None

        # Fallback title text if image isn't available
        if self.header_label is None:
            title_label = ttk.Label(
                main_frame, text="Invoice Extractor",
                font=('Segoe UI', 18, 'bold')
            )
            title_label.pack(pady=(0, 5))

        # Subtitle (two lines, with emphasis on the second line)
        subtitle_line1 = ttk.Label(
            main_frame,
            text="Download Invoices from Gmail, Parse them, export to Excel, Validate with SkuNexus,",
            font=('Segoe UI', 9),
            justify='center',
            anchor='center'
        )
        subtitle_line1.pack(pady=(0, 2))

        def _create_spaced_text(parent, text, font_obj, spacing_px=1):
            style = ttk.Style()
            bg = style.lookup('TFrame', 'background') or self.root.cget('bg')
            fg = style.lookup('TLabel', 'foreground') or 'black'
            canvas = tk.Canvas(parent, highlightthickness=0, bd=0, bg=bg)
            x = 0
            for ch in text:
                canvas.create_text(x, 0, text=ch, font=font_obj, anchor='nw', fill=fg)
                x += font_obj.measure(ch) + spacing_px
            width = max(1, x - spacing_px)
            height = font_obj.metrics('linespace')
            canvas.configure(width=width, height=height)
            return canvas

        subtitle_line2_font = tkfont.Font(family='Segoe UI', size=11)
        subtitle_line2 = _create_spaced_text(
            main_frame,
            "Burn and Churn Invoices like a Badass!",
            subtitle_line2_font,
            spacing_px=1
        )
        subtitle_line2.pack(pady=(0, 10))

        # Info frame
        info_frame = ttk.LabelFrame(main_frame, text="Status", padding=8)
        info_frame.pack(fill=tk.X, pady=(0, 10))

        # Credential status
        cred_exists = os.path.exists(os.path.join(self.required_dir, 'client_secret.json'))
        token_exists = os.path.exists(os.path.join(self.required_dir, 'token.pickle'))
        ocr_status = "Available" if OCR_AVAILABLE else "Not available (scanned PDFs will be skipped)"

        self.cred_label = ttk.Label(
            info_frame,
            text=(
                f"Credentials: {'Found' if cred_exists else 'MISSING - place client_secret.json in App/required'}"
            ),
            foreground='green' if cred_exists else 'red'
        )
        self.cred_label.pack(anchor=tk.W)

        self.auth_label = ttk.Label(
            info_frame,
            text=f"Authentication: {'Cached (token.pickle found)' if token_exists else 'Will authenticate on first run'}",
            foreground='green' if token_exists else 'gray'
        )
        self.auth_label.pack(anchor=tk.W)

        ttk.Label(
            info_frame,
            text=f"OCR (Tesseract): {ocr_status}",
            foreground='green' if OCR_AVAILABLE else 'orange'
        ).pack(anchor=tk.W)

        # Count existing invoices
        invoice_count = 0
        if os.path.exists(self.invoices_dir):
            invoice_count = len([
                f for f in os.listdir(self.invoices_dir)
                if f.lower().endswith(('.pdf', '.png', '.jpg', '.jpeg', '.tiff'))
            ])
        ttk.Label(
            info_frame,
            text=f"Invoices downloaded: {invoice_count}"
        ).pack(anchor=tk.W)

        # Date filter frame (inside Status)
        filter_frame = ttk.LabelFrame(info_frame, text="Filter Invoices by Date (Optional)", padding=8)
        filter_frame.pack(fill=tk.X, pady=(8, 0))

        self.date_filter_var = tk.BooleanVar(value=False)
        self.today_filter_var = tk.BooleanVar(value=False)
        self.date_from_var = tk.StringVar()
        self.date_to_var = tk.StringVar()

        self.date_filter_check = ttk.Checkbutton(
            filter_frame,
            text="Filter by date range",
            variable=self.date_filter_var,
            command=self._on_date_filter_toggle
        )
        self.date_filter_check.grid(row=0, column=0, sticky='w')

        ttk.Label(filter_frame, text="From (YYYY/MM/DD)").grid(row=0, column=1, padx=(10, 2))
        self.date_from_entry = ttk.Entry(filter_frame, textvariable=self.date_from_var, width=12)
        self.date_from_entry.grid(row=0, column=2, padx=(0, 10))

        ttk.Label(filter_frame, text="To (YYYY/MM/DD)").grid(row=0, column=3, padx=(0, 2))
        self.date_to_entry = ttk.Entry(filter_frame, textvariable=self.date_to_var, width=12)
        self.date_to_entry.grid(row=0, column=4)
        self.date_from_entry.bind(
            "<FocusIn>", lambda _e: self._open_date_picker_at(
                self.date_from_var, self.date_from_entry, "Select FROM Date"
            )
        )
        self.date_to_entry.bind(
            "<FocusIn>", lambda _e: self._open_date_picker_at(
                self.date_to_var, self.date_to_entry, "Select TO Date"
            )
        )

        self.today_filter_check = ttk.Checkbutton(
            filter_frame,
            text="All from Today",
            variable=self.today_filter_var,
            command=self._on_today_filter_toggle
        )
        self.today_filter_check.grid(row=1, column=0, sticky='w', pady=(4, 0))

        ttk.Label(
            filter_frame,
            text="Filtering by date may download already downloaded invoices.",
            foreground='orange'
        ).grid(row=2, column=0, columnspan=5, sticky='w', pady=(4, 0))

        # Buttons frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))

        self.go_button = ttk.Button(
            btn_frame, text="Go", command=self.start_processing,
            style='Accent.TButton'
        )
        self.go_button.pack(side=tk.LEFT, padx=(0, 5))

        self.stop_button = ttk.Button(
            btn_frame, text="Stop", command=self.stop_processing,
            state=tk.DISABLED
        )
        self.stop_button.pack(side=tk.LEFT, padx=(0, 5))

        ttk.Label(btn_frame, text="-").pack(side=tk.LEFT, padx=(0, 5))

        self.validate_button = ttk.Button(
            btn_frame, text="Validate POs", command=self.start_validation
        )
        self.validate_button.pack(side=tk.LEFT)

        # Batch export buttons (below main actions)
        batch_frame = ttk.Frame(main_frame)
        batch_frame.pack(fill=tk.X, pady=(0, 10))

        self.export_batches_button = ttk.Button(
            batch_frame, text="Export CSV Batches", command=self.export_csv_batches,
            state=tk.DISABLED
        )
        self.export_batches_button.pack(side=tk.LEFT, padx=(0, 5))

        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            main_frame, variable=self.progress_var,
            maximum=100, mode='determinate'
        )
        self.progress_bar.pack(fill=tk.X, pady=(0, 5))

        self.progress_label = ttk.Label(main_frame, text="Ready", font=('Segoe UI', 9))
        self.progress_label.pack(anchor=tk.W, pady=(0, 5))

        # Status log
        self.log_frame = ttk.LabelFrame(main_frame, text="Log", padding=5)
        self.log_frame.pack(fill=tk.BOTH, expand=True)
        self.log_frame.pack_propagate(False)
        self.log_frame.configure(height=200)

        self.log_text = tk.Text(
            self.log_frame, wrap=tk.WORD, font=('Consolas', 9),
            state=tk.DISABLED, bg='#1e1e1e', fg='#cccccc',
            insertbackground='white'
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(self.log_frame, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)

        # Configure text tags for colored output
        self.log_text.tag_configure('success', foreground='#4ec94e')
        self.log_text.tag_configure('error', foreground='#ff5555')
        self.log_text.tag_configure('warning', foreground='#ffaa00')

        # Initialize filter state and batch buttons
        self._update_date_filter_state()
        # Initialize batch buttons based on existing master files
        self._refresh_batch_buttons()
        self.log_text.tag_configure('info', foreground='#5599ff')
        self._ensure_log_min_height(200)

    def _init_header_source(self):
        if self.header_src_image is not None or not self.header_path or not os.path.exists(self.header_path):
            return
        try:
            if PIL_AVAILABLE:
                self.header_src_image = Image.open(self.header_path)
                self.header_src_width = self.header_src_image.width
            else:
                self.header_src_image = tk.PhotoImage(file=self.header_path)
                self.header_src_width = self.header_src_image.width()
        except Exception:
            self.header_src_image = None
            self.header_src_width = None

    def _load_header_image(self, target_width):
        self._init_header_source()
        if self.header_src_image is None or not self.header_src_width:
            return None

        target_width = int(target_width) if target_width else self.header_src_width
        if target_width <= 0:
            target_width = self.header_src_width

        scale = min(1.0, target_width / self.header_src_width)

        if PIL_AVAILABLE and isinstance(self.header_src_image, Image.Image):
            img = self.header_src_image
            if scale < 0.999:
                new_size = (
                    max(1, int(round(img.width * scale))),
                    max(1, int(round(img.height * scale)))
                )
                img = img.resize(new_size, Image.LANCZOS)
            return ImageTk.PhotoImage(img)

        img = self.header_src_image
        if scale >= 0.999:
            return img
        denom = 100
        num = max(1, int(round(scale * denom)))
        return img.zoom(num, num).subsample(denom, denom)

    def _update_header_width(self, target_width):
        if not self.header_label:
            return
        img = self._load_header_image(target_width)
        if not img:
            return
        self.header_image = img
        self.header_label.configure(image=self.header_image)
        try:
            self.header_current_width = self.header_image.width()
        except Exception:
            self.header_current_width = target_width

    def _ensure_log_min_height(self, min_height):
        def _apply():
            if not self.log_frame:
                return
            self.root.update_idletasks()
            current = self.log_frame.winfo_height()
            if current < min_height:
                width = self.root.winfo_width()
                height = self.root.winfo_height()
                delta = min_height - current
                self.root.geometry(f"{width}x{height + delta}")
                self.root.update_idletasks()
                self.log_frame.configure(height=min_height)
                self.log_frame.pack_propagate(False)
        self.root.after(0, _apply)

    def _ease_in_out_cubic(self, t):
        if t < 0.5:
            return 4 * t * t * t
        return 1 - pow(-2 * t + 2, 3) / 2

    def _animate_header_shrink(self):
        if not self.header_label or not self.header_base_width:
            return
        if self.header_animating or self.header_shrunken:
            return
        start_width = self.header_current_width or self.header_base_width
        target_width = int(round(self.header_base_width * 0.75))
        if target_width >= start_width:
            return

        self.header_animating = True
        duration_ms = 1750
        start_time = time.perf_counter()
        token = object()
        self._header_anim_token = token

        def step():
            if self._header_anim_token is not token:
                return
            elapsed_ms = (time.perf_counter() - start_time) * 1000.0
            t = min(1.0, elapsed_ms / duration_ms)
            eased = self._ease_in_out_cubic(t)
            width = int(round(start_width + (target_width - start_width) * eased))
            self._update_header_width(max(1, width))
            if t < 1.0:
                self.root.after(16, step)
            else:
                self.header_animating = False
                self.header_shrunken = True
                self.header_current_width = width

        step()

    def log(self, message, tag=None):
        """Add a message to the status log (thread-safe)."""
        def _update():
            self.log_text.config(state=tk.NORMAL)
            timestamp = datetime.now().strftime('%H:%M:%S')
            line = f"[{timestamp}] {message}\n"
            if tag:
                self.log_text.insert(tk.END, line, tag)
            else:
                self.log_text.insert(tk.END, line)
            self.log_text.see(tk.END)
            self.log_text.config(state=tk.DISABLED)

        self.root.after(0, _update)

    def set_progress(self, value, label_text=None):
        """Update progress bar and label (thread-safe)."""
        def _update():
            self.progress_var.set(value)
            if label_text:
                self.progress_label.config(text=label_text)
        self.root.after(0, _update)

    def start_processing(self):
        """Start the full extraction pipeline in a background thread."""
        if self.is_running:
            return

        self._animate_header_shrink()

        # Check for credentials
        if not os.path.exists(os.path.join(self.required_dir, 'client_secret.json')):
            self.log("ERROR: client_secret.json not found!", "error")
            self.log("Place your Google OAuth credentials file in App/required.", "error")
            return

        self.is_running = True
        self.go_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self._refresh_update_button_state()

        # Clear log
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

        # Choose a fresh output file + invoices folder every run
        self.output_file, self.invoices_dir = self._get_next_run_paths()
        self.log(f"Output file: {os.path.basename(self.output_file)}", "info")
        self.log(f"Invoices folder: {os.path.basename(self.invoices_dir)}", "info")

        self.set_progress(0, "Starting...")

        thread = threading.Thread(target=self.run_pipeline, daemon=True)
        thread.start()

    def stop_processing(self):
        """Signal the pipeline to stop."""
        self.is_running = False
        self.log("Stop requested - will halt after current operation...", "warning")

    def run_pipeline(self):
        """Main processing pipeline - runs in background thread."""
        try:
            # Phase 1: Gmail - download attachments
            self.set_progress(5, "Connecting to Gmail...")
            self.log("=== Phase 1: Fetching emails from Gmail ===", "info")

            client = GmailClient(
                self.base_dir,
                status_callback=self.log,
                data_dir=self.required_dir,
                invoices_dir=self.invoices_dir,
                expected_email=AUTHORIZED_GMAIL_ACCOUNT
            )
            client.authenticate()
            drive_client = DriveHistoryClient(client.creds, status_callback=self.log)

            if not self.is_running:
                self.finish("Stopped by user.")
                return

            self.set_progress(10, "Downloading attachments...")
            query, mode = self._build_gmail_query()
            if query is None:
                self.finish("Failed - invalid date filter.")
                return
            if mode == "label":
                self.log(f"Using Gmail label filter (skipping '{PROCESSED_LABEL_NAME}' emails).", "info")
            elif mode == "today":
                self.log("Downloading all emails from today (label filter ignored).", "warning")
            elif mode == "range":
                self.log("Downloading emails in date range (label filter ignored).", "warning")

            downloaded_attachments, total_emails, new_emails = (
                client.fetch_and_download_new_attachments(query=query)
            )

            if not self.is_running:
                self.finish("Stopped by user.")
                return

            # Phase 2: Parse invoices
            self.set_progress(40, "Parsing invoices...")
            self.log("", None)
            self.log("=== Phase 2: Parsing invoice files ===", "info")

            def _history_key(bill_no, po_number, vendor, invoice_date):
                parts = [
                    str(bill_no or '').strip(),
                    str(po_number or '').strip(),
                    str(vendor or '').strip(),
                    str(invoice_date or '').strip(),
                ]
                if not any(parts):
                    return ''
                return '|'.join([p.lower() for p in parts])

            history_rows = self._load_invoice_history(drive_client=drive_client)
            history_by_po = {}
            history_by_bill = {}
            history_keys = set()
            for row in history_rows:
                po = str(row.get('po_number', '')).strip()
                if po:
                    history_by_po.setdefault(po, []).append(row)
                bill_no_hist = str(row.get('bill_no', '')).strip()
                if bill_no_hist:
                    history_by_bill.setdefault(bill_no_hist, []).append(row)
                key = _history_key(
                    row.get('bill_no', ''),
                    po,
                    row.get('vendor', ''),
                    row.get('invoice_date', '')
                )
                if key:
                    history_keys.add(key)

            new_history_entries = []
            new_history_keys = set()
            folder_name = os.path.basename(self.invoices_dir)
            root_name = os.path.basename(os.path.dirname(self.invoices_dir))

            def _source_file(filename):
                return os.path.join(root_name, folder_name, filename).replace('\\', '/')

            sender_metadata = self._load_sender_metadata()
            sender_metadata_updated = False
            timestamp_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            for attachment in downloaded_attachments:
                filename = str((attachment or {}).get('filename', '')).strip()
                if not filename:
                    continue
                source_file = _source_file(filename)
                sender_metadata[source_file] = {
                    'source_file': source_file,
                    'filename': filename,
                    'sender_email': str((attachment or {}).get('sender_email', '') or '').strip().lower(),
                    'sender_header': str((attachment or {}).get('sender_header', '') or '').strip(),
                    'subject': str((attachment or {}).get('subject', '') or '').strip(),
                    'message_id': str((attachment or {}).get('message_id', '') or '').strip(),
                    'downloaded_at': timestamp_now,
                }
                sender_metadata_updated = True
            if sender_metadata_updated:
                self._save_sender_metadata(sender_metadata.values())
                self.log(
                    f"Stored sender metadata for {len(downloaded_attachments)} attachment(s)."
                )

            # Find all invoice files to parse
            all_invoice_files = []
            if os.path.exists(self.invoices_dir):
                for f in os.listdir(self.invoices_dir):
                    if f.lower().endswith(('.pdf', '.png', '.jpg', '.jpeg', '.tiff')):
                        all_invoice_files.append(f)

            if not all_invoice_files:
                self.log("No invoice files to parse.", "success")
            else:
                self.log(f"Found {len(all_invoice_files)} invoice files to parse.")

                success_count = 0
                error_count = 0
                error_files = []

                for i, filename in enumerate(all_invoice_files):
                    if not self.is_running:
                        self.finish("Stopped by user.")
                        return

                    progress = 40 + (50 * (i + 1) / len(all_invoice_files))
                    self.set_progress(
                        progress,
                        f"Parsing invoice {i + 1}/{len(all_invoice_files)}..."
                    )

                    filepath = os.path.join(self.invoices_dir, filename)
                    self.log(f"Processing: {filename}")

                    try:
                        source_path = _source_file(filename)
                        sender_entry = sender_metadata.get(source_path, {})
                        invoice_data = parse_invoice(
                            filepath,
                            self.log,
                            sender_email=sender_entry.get('sender_email', ''),
                            sender_header=sender_entry.get('sender_header', ''),
                        )

                        if invoice_data and invoice_data.get('not_an_invoice'):
                            write_not_invoice_row(self.output_file, source_path, self.log)
                            success_count += 1
                        elif invoice_data:
                            invoice_data['source_path'] = source_path
                            write_invoice_to_spreadsheet(
                                self.output_file, invoice_data, self.log
                            )
                            success_count += 1
                            bill_no = str(invoice_data.get('invoice_number', '')).strip()
                            po_number = str(invoice_data.get('po_number', '')).strip()
                            vendor = str(invoice_data.get('vendor', '')).strip()
                            invoice_date = str(invoice_data.get('date', '')).strip()
                            key = _history_key(bill_no, po_number, vendor, invoice_date)
                            if key and key not in history_keys and key not in new_history_keys:
                                new_history_entries.append({
                                    'bill_no': bill_no,
                                    'po_number': po_number,
                                    'vendor': vendor,
                                    'invoice_date': invoice_date,
                                    'downloaded_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                    'source_file': _source_file(filename),
                                })
                                new_history_keys.add(key)
                        else:
                            error_count += 1
                            error_files.append(filename)

                    except Exception as e:
                        self.log(f"  Failed to parse {filename}: {e}", "error")
                        error_count += 1
                        error_files.append(filename)

                # Apply duplicate markers and update history log
                if os.path.exists(self.output_file):
                    dup_summary = self._apply_duplicate_flags(
                        self.output_file,
                        history_by_po,
                        history_by_bill
                    )
                    dup_invoices = int((dup_summary or {}).get('duplicate_invoices', 0))
                    dup_rows = int((dup_summary or {}).get('duplicate_rows', 0))
                    if dup_invoices:
                        self.log(
                            f"Duplicate invoices flagged: {dup_invoices} invoice(s), {dup_rows} row(s).",
                            "warning"
                        )
                    else:
                        self.log("Duplicate invoices flagged: 0.")
                if new_history_entries:
                    self._append_invoice_history(new_history_entries, drive_client=drive_client)

                # Summary
                self.log("")
                self.log("=== Summary ===", "info")
                self.log(f"Emails checked: {total_emails} total, {new_emails} new")
                self.log(f"Attachments downloaded: {len(downloaded_attachments)}")
                self.log(f"Invoices parsed successfully: {success_count}", "success")
                if error_count:
                    self.log(f"Invoices with errors: {error_count}", "error")
                    for ef in error_files:
                        self.log(f"  - {ef}", "error")

            self.finish("Processing complete!")

        except FileNotFoundError as e:
            self.log(f"File not found: {e}", "error")
            self.finish("Failed - missing file.")
        except WrongAuthorizedAccountError as e:
            self.log(f"Authentication blocked: {e}", "error")
            self.log(
                f"Sign in with {AUTHORIZED_GMAIL_ACCOUNT} and try again.",
                "warning"
            )
            self.finish("Failed - wrong Gmail account.")
        except Exception as e:
            self.log(f"Unexpected error: {e}", "error")
            self.finish("Failed with error.")

    def finish(self, message):
        """Reset UI state after pipeline completes."""
        def _update():
            self.is_running = False
            self.go_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
            self._refresh_update_button_state()
            self.set_progress(100, message)
            self.log(message, "info")
            self._refresh_batch_buttons()
        self.root.after(0, _update)

    def open_spreadsheet(self):
        """Open the output spreadsheet in the default application."""
        if os.path.exists(self.output_file):
            os.startfile(self.output_file)
        else:
            self.log("Spreadsheet not found - run extraction first.", "warning")

    def open_invoices_folder(self):
        """Open the invoices folder in Explorer."""
        os.makedirs(self.invoices_dir, exist_ok=True)
        os.startfile(self.invoices_dir)

    def export_csv_batches(self):
        """Export the latest master XLSX into CSV batches capped by BATCH_ROW_LIMIT."""
        master_path = self._select_master_for_batching()
        if not master_path or not os.path.exists(master_path):
            self.log("No master spreadsheet found to export.", "warning")
            self._refresh_batch_buttons()
            return

        self.log(f"Exporting CSV batches from {os.path.basename(master_path)}...", "info")
        self.log(
            f"Batch limit: {BATCH_TOTAL_LINE_LIMIT} total lines per file "
            f"(header + up to {BATCH_ROW_LIMIT} data rows).",
            "info"
        )
        today_tag = f"{datetime.now().month}-{datetime.now().day}"
        if f"Invoices_Master_{today_tag}" not in os.path.basename(master_path):
            self.log(
                "Note: using the most recent master file (not today's date).",
                "warning"
            )

        raw_rows = read_spreadsheet_rows(master_path)
        if not raw_rows:
            self.log("Master spreadsheet has no data rows.", "warning")
            return

        def _is_total_amount_summary_row(row):
            # Keep master rows intact, but drop invoice summary rows for CSV batch uploads.
            ps = str(row.get('product_service', '')).strip().lower()
            if ps != 'total amount':
                return False
            return (
                str(row.get('sku', '')).strip() == ''
                and str(row.get('qty', '')).strip() == ''
                and str(row.get('rate', '')).strip() == ''
            )

        rows = [r for r in raw_rows if not _is_total_amount_summary_row(r)]
        dropped_summary_rows = len(raw_rows) - len(rows)
        if dropped_summary_rows:
            self.log(
                f"Removed {dropped_summary_rows} Total Amount summary row(s) from CSV batches.",
                "info"
            )

        if not rows:
            self.log("No exportable rows after removing summary rows.", "warning")
            return

        # Group rows by invoice (keep invoices intact)
        invoices = []
        current = []
        current_bill = None
        for row in rows:
            bill_no = str(row.get('bill_no', '')).strip()
            if bill_no:
                if current and bill_no != current_bill:
                    invoices.append(current)
                    current = []
                current_bill = bill_no
            if current_bill is None and not bill_no:
                current_bill = ''
            current.append(row)
        if current:
            invoices.append(current)

        # Build batches without splitting invoices
        batches = []
        batch_rows = 0
        current_batch = []
        for inv_rows in invoices:
            inv_count = len(inv_rows)
            if batch_rows and (batch_rows + inv_count) > BATCH_ROW_LIMIT:
                batches.append(current_batch)
                current_batch = []
                batch_rows = 0

            if inv_count > BATCH_ROW_LIMIT and batch_rows == 0:
                # Oversized invoice: put in its own batch
                batches.append(inv_rows)
                continue

            current_batch.extend(inv_rows)
            batch_rows += inv_count

        if current_batch:
            batches.append(current_batch)

        date_tag = _extract_date_tag_from_filename(os.path.basename(master_path))
        if not date_tag:
            date_tag = f"{datetime.now().month}-{datetime.now().day}"

        batches_dir = self._get_next_batches_dir(date_tag)
        os.makedirs(batches_dir, exist_ok=True)
        self.last_batches_dir = batches_dir

        headers = [header for _, header in COLUMNS]
        for idx, batch in enumerate(batches, start=1):
            if len(batches) == 1:
                filename = f"{BATCH_FOLDER_PREFIX}{date_tag}.csv"
            else:
                filename = f"{BATCH_FOLDER_PREFIX}{date_tag}_{idx}.csv"
            out_path = os.path.join(batches_dir, filename)
            with open(out_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for row in batch:
                    writer.writerow([row.get(key, '') for key, _ in COLUMNS])

        # Warn if any invoice exceeded the batch limit
        oversized = [len(inv) for inv in invoices if len(inv) > BATCH_ROW_LIMIT]
        if oversized:
            self.log(
                f"Warning: {len(oversized)} invoice(s) exceeded the batch limit "
                f"({BATCH_ROW_LIMIT} rows) and were exported alone.",
                "warning"
            )

        self.log(
            f"Exported {len(batches)} batch file(s) to {os.path.basename(batches_dir)}",
            "success"
        )
        self._refresh_batch_buttons()

    # Removed open_batches_folder button per UI simplification request.

    def start_validation(self):
        """Start the PO validation pipeline in a background thread."""
        if self.is_running:
            return

        # Check if any output XLSX files exist
        output_files = self._get_output_files_for_validation()
        if not output_files:
            self.log("ERROR: No output XLSX files found - run extraction first!", "error")
            return

        self.is_running = True
        self.go_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.validate_button.config(state=tk.DISABLED)
        self._refresh_update_button_state()

        # Clear log
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

        self.set_progress(0, "Starting validation...")

        thread = threading.Thread(target=self.run_validation_pipeline, daemon=True)
        thread.start()

    def _find_config_path(self, filename):
        candidate = os.path.join(self.required_dir, filename)
        if os.path.exists(candidate):
            return candidate
        return None

    def _resolve_row_memo(self, row, memo_by_bill):
        memo = str(row.get('memo', '')).strip()
        if memo:
            return memo
        bill_no = str(row.get('bill_no', '')).strip()
        if not bill_no:
            return ''
        return str(memo_by_bill.get(bill_no, '')).strip()

    def _build_shopify_core_updates(self, rows, memo_by_bill, po_cache, shopify_client):
        updates = {}
        stats = {
            'checked': 0,
            'matched': 0,
            'missing': 0,
            'mismatch': 0,
        }
        related_order_cache = {}
        bill_group_amount_pool = {}

        for row in rows:
            if not _is_core_row(row):
                continue

            row_num = row['_row_num']
            bill_no = str(row.get('bill_no', '')).strip()
            memo = self._resolve_row_memo(row, memo_by_bill)
            stats['checked'] += 1

            if not memo:
                updates[row_num] = ('NONE', 'missing')
                stats['missing'] += 1
                continue

            cache_entry = po_cache.get(memo, {})
            related_order_numbers = list(cache_entry.get('related_order_numbers') or [])
            related_key = tuple(related_order_numbers)

            if not related_order_numbers:
                self.log(
                    f"  No SkuNexus related order number found for PO {memo}; "
                    "Shopify CORE cannot be confirmed.",
                    "warning"
                )
                updates[row_num] = ('NONE', 'missing')
                stats['missing'] += 1
                continue

            if related_key not in related_order_cache:
                aggregated_amounts = []
                total_shopify_orders = 0
                related_label = ', '.join(related_order_numbers)
                self.log(f"Checking Shopify CORE using SkuNexus related order number(s): {related_label}...")

                for related_number in related_order_numbers:
                    result, error = shopify_client.get_order_number_core_amounts(related_number)
                    if error:
                        self.log(
                            f"  Shopify lookup failed for related order {related_number}: {error}",
                            "warning"
                        )
                        continue

                    orders = result.get('orders', [])
                    core_amounts = list(result.get('core_amounts', []))
                    total_shopify_orders += len(orders)
                    aggregated_amounts.extend(core_amounts)

                related_order_cache[related_key] = {
                    'core_amounts': aggregated_amounts,
                    'order_count': total_shopify_orders,
                }

                if aggregated_amounts:
                    sample = ', '.join(f"{a:.2f}" for a in aggregated_amounts[:6])
                    if len(aggregated_amounts) > 6:
                        sample += ", ..."
                    self.log(
                        f"  Found Shopify CORE amount(s): {sample} "
                        f"across {total_shopify_orders} order(s)"
                    )
                else:
                    self.log(
                        f"  No Shopify CORE line item found for related order number(s): "
                        f"{related_label}",
                        "warning"
                    )

            group_key = (bill_no, memo, related_key)
            if group_key not in bill_group_amount_pool:
                bill_group_amount_pool[group_key] = list(
                    related_order_cache.get(related_key, {}).get('core_amounts', [])
                )
            core_pool = bill_group_amount_pool[group_key]

            invoice_rate = _to_float_value(row.get('rate', ''))
            if not core_pool:
                updates[row_num] = ('NONE', 'missing')
                stats['missing'] += 1
                continue

            matched_idx = None
            if invoice_rate is not None:
                for idx, core_amount in enumerate(core_pool):
                    if abs(core_amount - invoice_rate) <= SHOPIFY_CORE_RATE_TOLERANCE:
                        matched_idx = idx
                        break

            if matched_idx is not None:
                matched_amount = core_pool.pop(matched_idx)
                updates[row_num] = (f"{matched_amount:.2f}", 'ok')
                stats['matched'] += 1
                continue

            if invoice_rate is None:
                selected_amount = core_pool.pop(0)
                updates[row_num] = (f"{selected_amount:.2f}", 'ok')
                stats['matched'] += 1
                continue

            closest_idx = min(range(len(core_pool)), key=lambda i: abs(core_pool[i] - invoice_rate))
            closest_amount = core_pool.pop(closest_idx)
            updates[row_num] = (f"{closest_amount:.2f}", 'mismatch')
            stats['mismatch'] += 1

        return updates, stats

    def run_validation_pipeline(self):
        """Validate POs against SkuNexus - runs in background thread."""
        try:
            self.log("=== SkuNexus PO Validation ===", "info")

            # Load SkuNexus credentials from config file
            config_path = self._find_config_path('skunexus_config.json')
            if not config_path or not os.path.exists(config_path):
                self.log("ERROR: skunexus_config.json not found!", "error")
                self.log(
                    "Create skunexus_config.json with 'email' and 'password' in App/required.",
                    "error"
                )
                self.finish_validation("Validation failed - missing config file")
                return

            with open(config_path, 'r') as f:
                skunexus_config = json.load(f)

            skunexus_email = skunexus_config.get('email', '')
            skunexus_password = skunexus_config.get('password', '')

            if not skunexus_email or not skunexus_password:
                self.log("ERROR: Invalid skunexus_config.json - missing email or password", "error")
                self.finish_validation("Validation failed - invalid config")
                return

            # Login to SkuNexus
            self.set_progress(5, "Logging into SkuNexus...")
            self.log("Connecting to SkuNexus...")

            client = SkuNexusClient(skunexus_email, skunexus_password)
            success, message = client.login()

            if not success:
                self.log(f"Failed to login to SkuNexus: {message}", "error")
                self.finish_validation("Validation failed - could not login to SkuNexus")
                return

            self.log("Successfully logged into SkuNexus", "success")

            # Shopify CORE validation is required for Validate POs.
            shopify_client = None
            shopify_enabled = False
            shopify_checked_count = 0
            shopify_matched_count = 0
            shopify_missing_count = 0
            shopify_mismatch_count = 0

            shopify_config_path = self._find_config_path('shopify_config.json')
            if not shopify_config_path or not os.path.exists(shopify_config_path):
                self.log("ERROR: shopify_config.json not found in App/required.", "error")
                self.finish_validation("Validation failed - missing Shopify config")
                return

            try:
                with open(shopify_config_path, 'r', encoding='utf-8-sig') as f:
                    shopify_config = json.load(f)
            except Exception as e:
                self.log(f"ERROR: Could not read shopify_config.json: {e}", "error")
                self.finish_validation("Validation failed - invalid Shopify config")
                return

            shop = (
                shopify_config.get('shop')
                or shopify_config.get('shop_domain')
                or shopify_config.get('store')
                or ''
            )
            client_id = shopify_config.get('client_id') or shopify_config.get('clientId') or ''
            client_secret = (
                shopify_config.get('client_secret')
                or shopify_config.get('clientSecret')
                or ''
            )
            scopes = shopify_config.get('scopes') or ['read_orders']
            if isinstance(scopes, str):
                scopes = [p.strip() for p in re.split(r'[,\s]+', scopes) if p.strip()]
            elif not isinstance(scopes, list):
                scopes = ['read_orders']
            if 'read_orders' not in scopes:
                scopes = list(scopes) + ['read_orders']
            api_version = shopify_config.get('api_version') or '2025-10'
            auth_mode = str(shopify_config.get('auth_mode') or 'auto').strip().lower()
            redirect_uri = str(shopify_config.get('redirect_uri') or '').strip()
            callback_host = str(shopify_config.get('callback_host') or '127.0.0.1').strip()
            callback_bind_host = str(
                shopify_config.get('callback_bind_host')
                or shopify_config.get('callback_host')
                or '0.0.0.0'
            ).strip()
            callback_port = int(shopify_config.get('callback_port') or 8765)
            callback_bind_port = int(
                shopify_config.get('callback_bind_port')
                or shopify_config.get('callback_port')
                or callback_port
            )
            callback_path = str(shopify_config.get('callback_path') or '/shopify/callback').strip()
            try:
                oauth_timeout = int(
                    shopify_config.get('oauth_timeout')
                    or shopify_config.get('auth_timeout_seconds')
                    or 120
                )
            except Exception:
                oauth_timeout = 120

            if not (shop and client_id and client_secret):
                self.log(
                    "ERROR: shopify_config.json is missing shop/client_id/client_secret.",
                    "error"
                )
                self.finish_validation("Validation failed - incomplete Shopify config")
                return

            self.set_progress(8, "Authenticating Shopify...")
            self.log(f"Authenticating Shopify (mode: {auth_mode or 'auto'})...")
            shopify_client = ShopifyClient(
                shop=shop,
                client_id=client_id,
                client_secret=client_secret,
                scopes=scopes,
                api_version=api_version,
                token_file=os.path.join(self.required_dir, 'shopify_token.json'),
                status_callback=self.log,
                auth_mode=auth_mode,
                callback_host=callback_host,
                redirect_uri=redirect_uri,
                callback_bind_host=callback_bind_host,
                callback_bind_port=callback_bind_port,
                callback_port=callback_port,
                callback_path=callback_path,
                oauth_timeout=oauth_timeout,
            )
            ok, shopify_message = shopify_client.authenticate()
            if not ok:
                self.log(f"ERROR: Shopify authentication failed: {shopify_message}", "error")
                self.finish_validation("Validation failed - Shopify authentication required")
                return
            self.log(f"Successfully connected to Shopify ({shopify_message})", "success")
            shopify_enabled = True

            if not self.is_running:
                self.finish_validation("Stopped by user.")
                return

            # Collect output files
            self.set_progress(10, "Reading output files...")
            output_files = self._get_output_files_for_validation()
            if not output_files:
                self.log("No output XLSX files found.", "warning")
                self.finish_validation("Validation complete - no data to validate")
                return

            vendor_aliases = load_vendor_aliases(self.app_dir)

            files_info = []
            total_rows = 0
            for filepath in output_files:
                rows = read_spreadsheet_rows(filepath)
                if rows:
                    files_info.append((filepath, rows))
                    total_rows += len(rows)

            if total_rows == 0:
                self.log("No data found in output XLSX files.", "warning")
                self.finish_validation("Validation complete - no data to validate")
                return

            self.log(f"Found {len(files_info)} output file(s) to validate")

            if not self.is_running:
                self.finish_validation("Stopped by user.")
                return

            # Validate rows across all files
            validated_count = 0
            passed_count = 0
            failed_count = 0
            not_found_count = 0
            skipped_count = 0
            already_validated_count = 0
            locked_files_count = 0
            po_cache = {}  # Cache SkuNexus data + margin by PO number

            processed_rows = 0

            def _pick_vendor(vendors):
                if not vendors:
                    return ''
                counts = {}
                for v in vendors:
                    counts[v] = counts.get(v, 0) + 1
                return max(counts, key=counts.get)

            for filepath, rows in files_info:
                if not self.is_running:
                    self.finish_validation("Stopped by user.")
                    return

                basename = os.path.basename(filepath)
                self.log("")
                self.log(f"--- Validating {basename} ---", "info")

                # Map Bill No. -> Memo (PO) so we can validate item rows that omit Memo
                memo_by_bill = {}
                for row in rows:
                    bill_no = str(row.get('bill_no', '')).strip()
                    memo = str(row.get('memo', '')).strip()
                    if bill_no and memo:
                        memo_by_bill[bill_no] = memo

                # Group rows by PO number (collect vendor/SKU hints)
                po_groups = {}
                for row in rows:
                    memo = str(row.get('memo', '')).strip()
                    if not memo:
                        bill_no = str(row.get('bill_no', '')).strip()
                        memo = str(memo_by_bill.get(bill_no, '')).strip()
                    if not memo:
                        continue
                    group = po_groups.setdefault(memo, {'rows': [], 'vendors': [], 'skus': []})
                    group['rows'].append(row)
                    vendor = str(row.get('vendor', '')).strip()
                    if vendor:
                        group['vendors'].append(vendor)
                    sku = _get_row_sku(row)
                    if sku:
                        group['skus'].append(sku)

                updates = {}
                margin_updates = {}
                first_margin_row_by_group = {}
                for row in rows:
                    if not self.is_running:
                        self.finish_validation("Stopped by user.")
                        return

                    processed_rows += 1
                    progress = 15 + (80 * (processed_rows / total_rows))
                    self.set_progress(progress, f"Validating row {processed_rows}/{total_rows}...")

                    row_num = row['_row_num']
                    bill_no = str(row.get('bill_no', '')).strip()

                    existing_validation = str(row.get('skunexus_validation', '')).strip()

                    memo = str(row.get('memo', '')).strip()
                    if not memo:
                        bill_no = str(row.get('bill_no', '')).strip()
                        memo = str(memo_by_bill.get(bill_no, '')).strip()
                    category = row.get('category', '')

                    # Skip rows without PO number
                    if not memo:
                        skipped_count += 1
                        continue

                    # Only validate SKU rows (Category/Account = Purchases)
                    if category != 'Purchases':
                        skipped_count += 1
                        continue

                    sku_value = _get_row_sku(row)
                    if not _looks_like_sku(sku_value):
                        skipped_count += 1
                        continue

                    # Get SkuNexus data (from cache or fetch)
                    if memo not in po_cache:
                        # Extract PO number without "PO" prefix
                        po_number = memo[2:] if memo.upper().startswith('PO') else memo

                        self.log(f"Fetching PO {po_number} from SkuNexus...")
                        group = po_groups.get(memo, {})
                        vendor_hint = _pick_vendor(group.get('vendors', []))
                        sku_hints = group.get('skus', [])
                        sn_data, error = client.get_best_po_with_line_items(
                            po_number,
                            invoice_vendor=vendor_hint,
                            invoice_skus=sku_hints,
                            vendor_aliases=vendor_aliases
                        )

                        if error:
                            self.log(f"  Could not find PO {po_number}: {error}", "warning")
                            po_cache[memo] = {
                                'sn_data': None,
                                'margin': None,
                                'related_order_numbers': [],
                            }
                        else:
                            margin_value, margin_error = client.get_po_margin(sn_data, po_number)
                            if margin_error:
                                self.log(f"  Margin unavailable for PO {po_number}: {margin_error}", "warning")
                            else:
                                self.log(f"  PO margin: {margin_value:.4f}")
                            related_order_numbers = _extract_related_order_numbers(sn_data)
                            if related_order_numbers:
                                self.log(
                                    "  Related order number(s): "
                                    f"{', '.join(related_order_numbers)}"
                                )
                            else:
                                self.log("  No related order number found in SkuNexus", "warning")
                            po_cache[memo] = {
                                'sn_data': sn_data,
                                'margin': margin_value,
                                'related_order_numbers': related_order_numbers,
                            }
                            self.log(f"  Found PO with {len(sn_data.get('lineItems', {}).get('rows', []))} line items")

                    cache_entry = po_cache.get(memo, {})
                    sn_data = cache_entry.get('sn_data')
                    margin_value = cache_entry.get('margin')

                    if margin_value is not None:
                        margin_group_key = (bill_no, memo)
                        if margin_group_key not in first_margin_row_by_group:
                            first_margin_row_by_group[margin_group_key] = row_num
                            margin_updates[row_num] = margin_value
                        else:
                            # Keep PO margin visible only on the first item row for this bill/PO.
                            margin_updates[row_num] = ''

                    if existing_validation:
                        already_validated_count += 1
                        continue

                    if sn_data is None:
                        # PO not found in SkuNexus
                        updates[row_num] = (False, ['PO not found in SkuNexus'])
                        validated_count += 1
                        not_found_count += 1
                        continue

                    # Validate this row against SkuNexus data
                    is_valid, failed_fields = validate_po_row(sn_data, row, vendor_aliases)
                    updates[row_num] = (is_valid, failed_fields)
                    validated_count += 1

                    if is_valid:
                        passed_count += 1
                    else:
                        failed_count += 1
                        sku = _get_row_sku(row) or 'N/A'
                        self.log(f"  Row {row_num} (SKU: {sku}) - FAILED: {', '.join(failed_fields)}", "warning")

                shopify_core_updates = {}
                if shopify_enabled and shopify_client:
                    (
                        shopify_core_updates,
                        shopify_stats
                    ) = self._build_shopify_core_updates(rows, memo_by_bill, po_cache, shopify_client)
                    shopify_checked_count += shopify_stats.get('checked', 0)
                    shopify_matched_count += shopify_stats.get('matched', 0)
                    shopify_missing_count += shopify_stats.get('missing', 0)
                    shopify_mismatch_count += shopify_stats.get('mismatch', 0)

                if updates or margin_updates or shopify_core_updates:
                    try:
                        write_validation_results(
                            filepath,
                            updates,
                            margin_updates,
                            shopify_core_updates
                        )
                        self.log(
                            f"Updated {len(updates)} validation row(s) and "
                            f"{len(margin_updates)} margin row(s) and "
                            f"{len(shopify_core_updates)} Shopify CORE row(s) in {basename}",
                            "success"
                        )
                    except PermissionError as e:
                        locked_files_count += 1
                        self.log(
                            f"Could not update {basename}: file is locked/open in Excel.",
                            "error"
                        )
                        self.log(f"  {e}", "warning")
                        self.log(
                            "  Close the workbook and run Validate POs again.",
                            "warning"
                        )
                else:
                    self.log(f"No rows needed validation in {basename}", "success")

            # Summary
            self.log("")
            self.log("=== Validation Summary ===", "info")
            self.log(f"Total rows validated: {validated_count}")
            self.log(f"Passed: {passed_count}", "success")
            if failed_count:
                self.log(f"Failed: {failed_count}", "error")
            if not_found_count:
                self.log(f"POs not found: {not_found_count}", "warning")
            if already_validated_count:
                self.log(f"Already validated rows skipped: {already_validated_count}")
            if skipped_count:
                self.log(f"Non-SKU/Non-PO rows skipped: {skipped_count}")
            if shopify_enabled:
                self.log(f"Shopify CORE rows checked: {shopify_checked_count}")
                self.log(f"Shopify CORE matched: {shopify_matched_count}", "success")
                if shopify_mismatch_count:
                    self.log(f"Shopify CORE mismatched: {shopify_mismatch_count}", "warning")
                if shopify_missing_count:
                    self.log(f"Shopify CORE missing: {shopify_missing_count}", "error")
            if locked_files_count:
                self.log(
                    f"Files skipped due to lock/open Excel window: {locked_files_count}",
                    "warning"
                )

            self.finish_validation("Validation complete!")

        except Exception as e:
            self.log(f"Validation error: {e}", "error")
            import traceback
            self.log(traceback.format_exc(), "error")
            self.finish_validation("Validation failed with error.")

    def finish_validation(self, message):
        """Reset UI state after validation completes."""
        def _update():
            self.is_running = False
            self.go_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
            self.validate_button.config(state=tk.NORMAL)
            self._refresh_update_button_state()
            self.set_progress(100, message)
            self.log(message, "info")
            self._refresh_batch_buttons()
        self.root.after(0, _update)


def main():
    if not _ensure_single_instance("Invoice Extractor"):
        return
    root = tk.Tk()

    # Try to use a modern theme
    try:
        style = ttk.Style()
        available_themes = style.theme_names()
        if 'vista' in available_themes:
            style.theme_use('vista')
        elif 'clam' in available_themes:
            style.theme_use('clam')
    except Exception:
        pass

    app = InvoiceExtractorGUI(root)
    root.mainloop()


if __name__ == '__main__':
    main()
